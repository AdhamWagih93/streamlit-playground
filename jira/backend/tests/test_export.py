"""Export endpoints: filtered issues (CSV/JSON/XLSX) and insights (JSON/CSV/MD).

Verifies content types, attachment headers, parseable content, format
validation, and that RBAC scoping is preserved.
"""
from __future__ import annotations

import csv
import io
import json
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient

from app.core.config import settings


@pytest.fixture(scope="module")
def client():
    from app.main import app

    with TestClient(app) as c:
        yield c


@pytest.fixture(scope="module")
def admin(client):
    r = client.post("/api/auth/login", data={
        "username": settings.bootstrap_admin_email,
        "password": settings.bootstrap_admin_password,
    })
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


@pytest.fixture(scope="module")
def project(client, admin):
    key = "X" + uuid4().hex[:5].upper()
    proj = client.post("/api/projects", headers=admin, json={"key": key, "name": f"Exp {key}"})
    assert proj.status_code == 201, proj.text
    pid = proj.json()["id"]
    story = next(t for t in client.get("/api/meta/issue-types", headers=admin).json()
                 if t["name"] == "Story")["id"]
    keys = []
    for i in range(4):
        r = client.post("/api/issues", headers=admin, json={
            "project_id": pid, "type_id": story, "summary": f"export issue {i}",
            "label_names": ["exp"]})
        assert r.status_code == 201
        keys.append(r.json()["key"])
    return {"id": pid, "key": key, "issue_keys": keys}


def _register(client, label):
    email = f"{label}-{uuid4().hex[:8]}@example.com"
    client.post("/api/auth/register", json={
        "username": f"{label}_{uuid4().hex[:6]}", "email": email,
        "display_name": label, "password": "password123"})
    tok = client.post("/api/auth/login", data={"username": email, "password": "password123"})
    return {"Authorization": f"Bearer {tok.json()['access_token']}"}


# --- Issue export ----------------------------------------------------------
def test_export_issues_csv(client, admin, project):
    r = client.get("/api/search/export", headers=admin, params={"tql": f"project = {project['key']}", "format": "csv"})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("text/csv")
    assert "attachment" in r.headers["content-disposition"]
    assert r.headers["content-disposition"].endswith('trackly-issues.csv"')
    rows = list(csv.DictReader(io.StringIO(r.content.decode("utf-8-sig"))))
    assert len(rows) == 4
    assert {row["Key"] for row in rows} == set(project["issue_keys"])
    assert rows[0]["Labels"] == "exp"


def test_export_issues_json(client, admin, project):
    r = client.get("/api/search/export", headers=admin, params={"tql": f"project = {project['key']}", "format": "json"})
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/json")
    data = json.loads(r.content)
    assert len(data) == 4 and {d["Key"] for d in data} == set(project["issue_keys"])


def test_export_issues_xlsx(client, admin, project):
    r = client.get("/api/search/export", headers=admin, params={"tql": f"project = {project['key']}", "format": "xlsx"})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content[:2] == b"PK"  # xlsx is a zip
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws.max_row == 5  # header + 4 issues
    assert ws.cell(row=1, column=1).value == "Key"


def test_export_issues_format_and_tql_validation(client, admin, project):
    assert client.get("/api/search/export", headers=admin, params={"format": "pdf"}).status_code == 400
    assert client.get("/api/search/export", headers=admin,
                      params={"tql": "project == (", "format": "csv"}).status_code == 400


def test_export_issues_respects_rbac(client, admin, project):
    """An outsider's export is scoped to browsable projects → empty."""
    O = _register(client, "exoutsider")
    r = client.get("/api/search/export", headers=O, params={"tql": f"project = {project['key']}", "format": "json"})
    assert r.status_code == 200
    assert json.loads(r.content) == []


# --- Insights export -------------------------------------------------------
def test_export_project_insights_formats(client, admin, project):
    j = client.get(f"/api/analytics/projects/{project['key']}/export", headers=admin, params={"format": "json"})
    assert j.status_code == 200 and j.headers["content-type"].startswith("application/json")
    body = json.loads(j.content)
    assert body["project_key"] == project["key"] and body["total_issues"] == 4

    c = client.get(f"/api/analytics/projects/{project['key']}/export", headers=admin, params={"format": "csv"})
    assert c.status_code == 200 and c.headers["content-type"].startswith("text/csv")
    assert "Project insights" in c.content.decode("utf-8-sig")

    m = client.get(f"/api/analytics/projects/{project['key']}/export", headers=admin, params={"format": "md"})
    assert m.status_code == 200 and m.headers["content-type"].startswith("text/markdown")
    assert m.content.decode().startswith("# Insights")

    assert client.get(f"/api/analytics/projects/{project['key']}/export",
                      headers=admin, params={"format": "pdf"}).status_code == 400


def test_export_insights_window_applies(client, admin, project):
    past = client.get(f"/api/analytics/projects/{project['key']}/export",
                      headers=admin, params={"format": "json", "to": "2000-01-01"})
    assert past.status_code == 200
    assert json.loads(past.content)["total_issues"] == 0


def test_export_overview_and_my(client, admin, project):
    ov = client.get("/api/analytics/overview/export", headers=admin, params={"format": "csv"})
    assert ov.status_code == 200 and "Instance insights" in ov.content.decode("utf-8-sig")
    mine = client.get("/api/analytics/my/export", headers=admin, params={"format": "json"})
    assert mine.status_code == 200 and json.loads(mine.content)["scope"] == "mine"


def test_export_insights_rbac(client, admin, project):
    O = _register(client, "exins")
    assert client.get(f"/api/analytics/projects/{project['key']}/export", headers=O,
                      params={"format": "json"}).status_code == 403
    assert client.get("/api/analytics/overview/export", headers=O,
                      params={"format": "json"}).status_code == 403
