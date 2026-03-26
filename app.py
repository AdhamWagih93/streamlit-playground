"""
Ollama Document Chatbot — Streamlit Page
Chat with your documents using local Ollama models.
Supports TXT, MD, PDF, and DOCX files.
Queue-based access: one user at a time.
Designed as a page within a multi-page Streamlit app.
"""

import base64
import json
import os
import re
import subprocess
import tempfile
import time
import hashlib
from datetime import datetime, timedelta, timezone
from pathlib import Path
from io import BytesIO
from typing import Optional

try:
    from zoneinfo import ZoneInfo
except ImportError:
    from backports.zoneinfo import ZoneInfo  # Python 3.8 fallback

import requests
import streamlit as st

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    Workbook = None

# ---------------------------------------------------------------------------
# Optional heavy imports
# ---------------------------------------------------------------------------
try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    from PyPDF2 import PdfReader
except ImportError:
    PdfReader = None

try:
    from docx import Document as DocxDocument
except ImportError:
    DocxDocument = None

try:
    import olefile
except ImportError:
    olefile = None

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    psycopg2 = None


# ---------------------------------------------------------------------------
# Configuration — set these once
# ---------------------------------------------------------------------------
OLLAMA_URL = "http://ef-nexus-03:8081"
MODEL = "qwen3.5:9b"
ENABLE_VISION = True                # set True to enable image upload (jpg, png, bmp)
VISION_MODEL = "qwen2.5-vl:7b"     # model with vision support (qwen3.5 is text-only)
ENABLE_DANGER_ZONE = False          # set True to show the "Clear All History" button in admin view
HISTORY_SCHEMA = "public"           # postgres schema
HISTORY_TABLE  = "chatbot_history"  # postgres table name

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
PROMPT_TIMEOUT_S     = 120   # active prompt expires after 2 min (stale detection)
QUEUE_WAIT_TIMEOUT_S = 300  # queued entries expire after 5 min
QUEUE_TABLE          = "prompt_queue"  # postgres table for shared queue
CHARS_PER_TOKEN_ESTIMATE = 4  # rough char-to-token ratio
LOCAL_TZ = ZoneInfo("Africa/Cairo")
MAX_PREVIEW_LINES = 12        # collapse messages longer than this
MAX_PREVIEW_CHARS = 500       # collapse messages longer than this

# ---------------------------------------------------------------------------
# Saved prompts — per team and per role
# Each entry: {"label": short button name, "prompt": full text}
# Users see prompts for all their teams (st.session_state.teams)
# and all their roles (st.session_state.user_roles).
# ---------------------------------------------------------------------------
TEAM_PROMPTS: dict[str, list[dict[str, str]]] = {
    "QC": [
        {
            "label": "Generate Test Cases",
            "prompt": (
                "You are an expert software testing analyst. "
                "Write software testing at least 200 test cases for the attached document "
                "for the positive and negative test cases in excel table format with columns: "
                "Test Case ID, Requirement, Description, Expected Results, Actual Results, "
                "Pass/Fail and Comments."
            ),
        },
        {
            "label": "Complete Remaining Test Cases",
            "prompt": (
                "You are an expert software testing analyst. "
                "Check the attached document, and the attached test cases, then write "
                "software testing for the remaining test cases — positive and negative — "
                "in excel table format with columns: Test Case ID, Requirement, Description, "
                "Expected Results, Actual Results, Pass/Fail and Comments."
            ),
        },
    ],
}

ROLE_PROMPTS: dict[str, list[dict[str, str]]] = {
    # Example:
    # "admin": [{"label": "System Report", "prompt": "Generate a system health report..."}],
}


# ---------------------------------------------------------------------------
# Page-scoped CSS
# ---------------------------------------------------------------------------
CUSTOM_CSS = """
<style>
:root {
    --bg-primary: #f5f2eb;
    --bg-secondary: #eae6dd;
    --bg-card: #ffffff;
    --accent: #4a90d9;
    --accent-hover: #3a7bc8;
    --accent-subtle: rgba(74, 144, 217, 0.08);
    --text-primary: #1a1a2e;
    --text-secondary: #6b7280;
    --text-muted: #9ca3af;
    --border: #d5d0c4;
    --border-light: #e8e4db;
    --success: #2d8a4e;
    --warning: #b8860b;
    --danger: #c0392b;
    --radius: 12px;
    --radius-sm: 8px;
}

/* ---------- Page background ---------- */
.stApp, section[data-testid="stMain"] {
    background-color: var(--bg-primary) !important;
}

.stChatMessage {
    border-radius: var(--radius) !important;
    border: 1px solid var(--border-light) !important;
    background: var(--bg-card) !important;
    margin-bottom: 0.6rem !important;
    padding: 1rem 1.25rem !important;
    position: relative;
}

.stButton > button {
    border-radius: var(--radius-sm) !important;
    font-weight: 500 !important;
    font-size: 0.85rem !important;
    transition: all 0.15s ease !important;
    border: 1px solid var(--border) !important;
    background: var(--bg-card) !important;
    color: var(--text-primary) !important;
}
.stButton > button:hover {
    border-color: var(--accent) !important;
    color: var(--accent) !important;
    background: var(--accent-subtle) !important;
}

.status-badge {
    display: inline-flex;
    align-items: center;
    gap: 5px;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 500;
    white-space: nowrap;
    vertical-align: middle;
}
.status-active {
    background: rgba(45, 138, 78, 0.1);
    color: var(--success);
    border: 1px solid rgba(45, 138, 78, 0.25);
}
.status-busy {
    background: rgba(184, 134, 11, 0.1);
    color: var(--warning);
    border: 1px solid rgba(184, 134, 11, 0.25);
}
.status-offline {
    background: rgba(192, 57, 43, 0.1);
    color: var(--danger);
    border: 1px solid rgba(192, 57, 43, 0.25);
}

.doc-chip {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius-sm);
    padding: 5px 11px;
    margin: 2px;
    font-size: 0.8rem;
    color: var(--text-primary);
}

.panel-section-title {
    font-size: 0.72rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    color: var(--text-secondary);
    margin-bottom: 0.35rem;
}

.divider {
    border: none;
    border-top: 1px solid var(--border-light);
    margin: 0.75rem 0;
}

.toolbar-brand {
    font-size: 1.05rem;
    font-weight: 700;
    color: var(--text-primary);
    letter-spacing: -0.02em;
    white-space: nowrap;
}
.toolbar-row {
    display: flex;
    align-items: center;
    gap: 10px;
}
.toolbar-meta {
    font-size: 0.78rem;
    color: var(--text-secondary);
    white-space: nowrap;
}

.empty-state {
    text-align: center;
    padding: 4rem 1rem 2rem;
    color: var(--text-secondary);
}
.empty-state h2 {
    font-size: 1.4rem;
    font-weight: 600;
    color: var(--text-primary);
    margin-bottom: 0.5rem;
}
.empty-state p {
    font-size: 0.92rem;
    max-width: 420px;
    margin: 0 auto;
    line-height: 1.6;
}

.lobby-card {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 2.5rem 2rem;
    text-align: center;
    max-width: 440px;
    margin: 3rem auto;
    box-shadow: 0 1px 3px rgba(0, 0, 0, 0.06);
}
.lobby-card h2 {
    font-size: 1.35rem;
    font-weight: 600;
    color: var(--text-primary);
    margin-bottom: 0.4rem;
}
.lobby-card p {
    color: var(--text-secondary);
    font-size: 0.9rem;
    line-height: 1.5;
    margin-bottom: 1.2rem;
}

div[data-testid="stPopover"] > div {
    min-width: 320px !important;
}

.doc-bar {
    display: flex;
    flex-wrap: wrap;
    align-items: center;
    gap: 6px;
    padding: 0.5rem 0;
}
.doc-bar-label {
    font-size: 0.75rem;
    font-weight: 600;
    color: var(--text-secondary);
    text-transform: uppercase;
    letter-spacing: 0.04em;
    margin-right: 4px;
}

/* ---------- Saved prompts ---------- */
.saved-prompts-bar {
    padding: 0.5rem 0;
    margin-bottom: 0.25rem;
}
.saved-prompts-label {
    font-size: 0.7rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    color: var(--text-muted);
    margin-bottom: 0.4rem;
}

/* ---------- Admin queue monitor ---------- */
.queue-monitor {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 1rem 1.25rem;
    margin-bottom: 1rem;
}
.queue-monitor-title {
    display: flex;
    align-items: center;
    gap: 8px;
    font-size: 0.82rem;
    font-weight: 600;
    color: var(--text-primary);
    margin-bottom: 0.65rem;
}
.queue-monitor-title .qm-dot {
    width: 8px; height: 8px;
    border-radius: 50%;
    background: var(--success);
    animation: qm-blink 1.5s ease-in-out infinite;
}
@keyframes qm-blink {
    0%, 100% { opacity: 1; }
    50% { opacity: 0.3; }
}
.queue-monitor-title .qm-dot-idle {
    background: var(--text-muted);
    animation: none;
}
.qm-entry {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 0.5rem 0.65rem;
    border-radius: var(--radius-sm);
    margin-bottom: 0.35rem;
    font-size: 0.8rem;
}
.qm-entry-active {
    background: rgba(74, 144, 217, 0.08);
    border: 1px solid rgba(74, 144, 217, 0.2);
}
.qm-entry-waiting {
    background: rgba(184, 134, 11, 0.06);
    border: 1px solid rgba(184, 134, 11, 0.15);
}
.qm-pos {
    font-weight: 700;
    font-size: 0.85rem;
    min-width: 24px;
    text-align: center;
}
.qm-pos-active { color: var(--accent); }
.qm-pos-waiting { color: var(--warning); }
.qm-user {
    font-weight: 600;
    color: var(--text-primary);
}
.qm-detail {
    color: var(--text-muted);
    font-size: 0.72rem;
    margin-left: auto;
    white-space: nowrap;
}

/* ---------- Prompt queue status ---------- */
.queue-card {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 1.5rem 1.75rem;
    text-align: center;
    margin: 1rem auto;
    max-width: 420px;
    box-shadow: 0 1px 4px rgba(0, 0, 0, 0.06);
    animation: queue-pulse 2s ease-in-out infinite;
}
@keyframes queue-pulse {
    0%, 100% { border-color: var(--border); }
    50% { border-color: var(--accent); }
}
.queue-card .queue-position {
    font-size: 2.2rem;
    font-weight: 700;
    color: var(--accent);
    line-height: 1.1;
}
.queue-card .queue-label {
    font-size: 0.85rem;
    color: var(--text-secondary);
    margin-top: 0.3rem;
}
.queue-card .queue-hint {
    font-size: 0.75rem;
    color: var(--text-muted);
    margin-top: 0.75rem;
}
.queue-bar {
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 10px;
    padding: 0.5rem 0.75rem;
    margin-bottom: 0.5rem;
    background: rgba(74, 144, 217, 0.06);
    border: 1px solid rgba(74, 144, 217, 0.15);
    border-radius: var(--radius-sm);
    font-size: 0.82rem;
    color: var(--accent);
    font-weight: 500;
}

/* ---------- Collapsible message ---------- */
.msg-truncated {
    position: relative;
}
.msg-truncated::after {
    content: "";
    position: absolute;
    bottom: 0;
    left: 0;
    right: 0;
    height: 2.5rem;
    background: linear-gradient(transparent, var(--bg-card));
    pointer-events: none;
    border-radius: 0 0 var(--radius-sm) var(--radius-sm);
}

/* Code snippet expanders */
.code-snippet-expander [data-testid="stExpander"] {
    border: 1px solid var(--border-light) !important;
    border-radius: var(--radius-sm) !important;
    background: var(--bg-primary) !important;
    margin: 0.35rem 0 !important;
}
.code-snippet-expander [data-testid="stExpander"] summary {
    font-size: 0.82rem !important;
    font-weight: 500 !important;
    color: var(--text-secondary) !important;
    padding: 0.45rem 0.75rem !important;
}
.code-snippet-expander [data-testid="stExpander"] summary:hover {
    color: var(--accent) !important;
}

/* ---------- Message metadata ---------- */
.msg-meta {
    display: flex;
    justify-content: flex-end;
    gap: 12px;
    margin-top: 0.4rem;
    font-size: 0.7rem;
    color: var(--text-muted);
    letter-spacing: 0.01em;
}
.msg-meta span {
    display: inline-flex;
    align-items: center;
    gap: 3px;
}

/* ---------- Document preview ---------- */
.doc-preview-card {
    background: var(--bg-primary);
    border: 1px solid var(--border-light);
    border-radius: var(--radius-sm);
    margin-top: 0.5rem;
}
.doc-preview-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 0.5rem 0.75rem;
    border-bottom: 1px solid var(--border-light);
}
.doc-preview-name {
    font-size: 0.8rem;
    font-weight: 600;
    color: var(--text-primary);
}
.doc-preview-stats {
    display: flex;
    gap: 10px;
    font-size: 0.7rem;
    color: var(--text-muted);
}
.doc-preview-stats span {
    display: inline-flex;
    align-items: center;
    gap: 3px;
}
.doc-preview-body {
    padding: 0.6rem 0.75rem;
    font-size: 0.78rem;
    color: var(--text-secondary);
    line-height: 1.55;
    max-height: 160px;
    overflow-y: auto;
    white-space: pre-wrap;
    word-break: break-word;
    font-family: 'SF Mono', 'Fira Code', 'Consolas', monospace;
}

/* ---------- History / Admin view ---------- */
.history-section {
    margin-top: 2.5rem;
    border-top: 2px solid var(--border);
    padding-top: 1.5rem;
}
.history-section-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 1.25rem;
}
.history-section-title {
    display: flex;
    align-items: center;
    gap: 10px;
}
.history-section-title h3 {
    font-size: 1.05rem;
    font-weight: 700;
    color: var(--text-primary);
    margin: 0;
    letter-spacing: -0.02em;
}
.admin-badge {
    display: inline-flex;
    align-items: center;
    padding: 2px 9px;
    border-radius: 20px;
    font-size: 0.67rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    background: rgba(74, 144, 217, 0.1);
    color: var(--accent);
    border: 1px solid rgba(74, 144, 217, 0.3);
}
.stat-grid {
    display: grid;
    grid-template-columns: repeat(6, 1fr);
    gap: 0.75rem;
    margin-bottom: 0.75rem;
}
.stat-card {
    background: var(--bg-card);
    border: 1px solid var(--border-light);
    border-radius: var(--radius);
    padding: 1rem 1rem 0.85rem;
    text-align: center;
    position: relative;
    overflow: hidden;
}
.stat-card::before {
    content: "";
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: var(--accent);
    opacity: 0.4;
    border-radius: var(--radius) var(--radius) 0 0;
}
.stat-card .stat-value {
    font-size: 1.55rem;
    font-weight: 700;
    color: var(--text-primary);
    line-height: 1.15;
}
.stat-card .stat-label {
    font-size: 0.7rem;
    font-weight: 500;
    color: var(--text-secondary);
    text-transform: uppercase;
    letter-spacing: 0.06em;
    margin-top: 0.25rem;
}
.stat-card .stat-sub {
    font-size: 0.68rem;
    color: var(--text-muted);
    margin-top: 0.15rem;
}
.stat-meta-row {
    display: flex;
    gap: 20px;
    font-size: 0.75rem;
    color: var(--text-muted);
    padding: 0.4rem 0 1rem;
    flex-wrap: wrap;
}
.stat-meta-row span strong {
    color: var(--text-secondary);
    font-weight: 600;
}
.filter-bar {
    background: var(--bg-card);
    border: 1px solid var(--border-light);
    border-radius: var(--radius);
    padding: 0.85rem 1rem;
    margin-bottom: 1rem;
}
.filter-bar-label {
    font-size: 0.7rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    color: var(--text-secondary);
    margin-bottom: 0.5rem;
}
.session-card {
    background: var(--bg-card);
    border: 1px solid var(--border-light);
    border-radius: var(--radius);
    margin-bottom: 0.6rem;
    overflow: hidden;
}
.session-card-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 0.7rem 1rem;
    background: var(--bg-secondary);
    border-bottom: 1px solid var(--border-light);
    flex-wrap: wrap;
    gap: 6px;
}
.session-card-id {
    font-family: 'SF Mono', 'Fira Code', 'Consolas', monospace;
    font-size: 0.72rem;
    color: var(--text-muted);
    background: var(--bg-primary);
    border: 1px solid var(--border-light);
    border-radius: 4px;
    padding: 1px 6px;
}
.session-card-user {
    font-size: 0.8rem;
    font-weight: 600;
    color: var(--text-primary);
}
.session-card-meta {
    display: flex;
    gap: 12px;
    font-size: 0.72rem;
    color: var(--text-muted);
    flex-wrap: wrap;
}
.history-row {
    background: var(--bg-card);
    border: 1px solid var(--border-light);
    border-left: 3px solid var(--border);
    border-radius: var(--radius-sm);
    padding: 0.75rem 1rem;
    margin-bottom: 0.35rem;
}
.history-row.role-user {
    border-left-color: var(--accent);
}
.history-row.role-assistant {
    border-left-color: var(--success);
}
.history-row-header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 0.3rem;
    flex-wrap: wrap;
    gap: 4px;
}
.history-role-badge {
    display: inline-flex;
    align-items: center;
    padding: 1px 8px;
    border-radius: 10px;
    font-size: 0.68rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}
.history-role-badge.role-user {
    background: rgba(74, 144, 217, 0.1);
    color: var(--accent);
    border: 1px solid rgba(74, 144, 217, 0.2);
}
.history-role-badge.role-assistant {
    background: rgba(45, 138, 78, 0.1);
    color: var(--success);
    border: 1px solid rgba(45, 138, 78, 0.2);
}
.history-row-ident {
    display: flex;
    gap: 8px;
    align-items: center;
    font-size: 0.72rem;
    color: var(--text-muted);
}
.history-content {
    font-size: 0.84rem;
    color: var(--text-primary);
    line-height: 1.5;
    margin-bottom: 0.3rem;
    word-break: break-word;
}
.history-row-meta {
    font-size: 0.7rem;
    color: var(--text-muted);
    display: flex;
    gap: 14px;
    flex-wrap: wrap;
}
.chart-card {
    background: var(--bg-card);
    border: 1px solid var(--border-light);
    border-radius: var(--radius);
    padding: 1rem 1.25rem 0.5rem;
    margin-bottom: 0.75rem;
}
.chart-card-title {
    font-size: 0.78rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    color: var(--text-secondary);
    margin-bottom: 0.5rem;
}
.chart-section-divider {
    display: flex;
    align-items: center;
    gap: 10px;
    margin: 1.25rem 0 0.75rem;
}
.chart-section-divider span {
    font-size: 0.78rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    color: var(--text-secondary);
    white-space: nowrap;
}
.chart-section-divider::after {
    content: "";
    flex: 1;
    height: 1px;
    background: var(--border-light);
}
.chart-card .stDataFrame {
    border: none !important;
}
.chart-card .stDataFrame [data-testid="stDataFrameResizable"] {
    border: 1px solid var(--border-light) !important;
    border-radius: var(--radius-sm) !important;
}

/* ---------- Code mode ---------- */
.code-output-panel {
    background: var(--bg-card);
    border: 2px solid var(--accent);
    border-radius: var(--radius);
    margin-top: 0.75rem;
    overflow: hidden;
}
.code-output-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 0.6rem 1rem;
    background: var(--accent-subtle);
    border-bottom: 1px solid rgba(74, 144, 217, 0.2);
}
.code-output-header span {
    font-size: 0.75rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    color: var(--accent);
}
.code-output-body {
    padding: 1rem;
}
.mode-pill {
    display: inline-flex;
    align-items: center;
    gap: 5px;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 0.72rem;
    font-weight: 600;
    letter-spacing: 0.04em;
}
.mode-pill.mode-chat {
    background: rgba(45, 138, 78, 0.1);
    color: var(--success);
    border: 1px solid rgba(45, 138, 78, 0.25);
}
.mode-pill.mode-code {
    background: rgba(74, 144, 217, 0.1);
    color: var(--accent);
    border: 1px solid rgba(74, 144, 217, 0.25);
}
</style>
"""

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def count_words(text: str) -> int:
    return len(text.split())


def estimate_tokens(text: str) -> int:
    return max(1, len(text) // CHARS_PER_TOKEN_ESTIMATE)


def format_number(n: int) -> str:
    if n >= 1_000_000:
        return f"{n / 1_000_000:.1f}M"
    if n >= 1_000:
        return f"{n / 1_000:.1f}k"
    return str(n)


def now_local() -> datetime:
    """Return current time in LOCAL_TZ (Africa/Cairo)."""
    return datetime.now(LOCAL_TZ)


def to_local(dt: datetime) -> datetime:
    """Convert a datetime to LOCAL_TZ. Handles naive (assumes UTC) and aware."""
    if dt is None:
        return dt
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(LOCAL_TZ)


def format_duration(seconds: float) -> str:
    if seconds < 1:
        return f"{seconds * 1000:.0f}ms"
    if seconds < 60:
        return f"{seconds:.1f}s"
    mins = int(seconds // 60)
    secs = seconds % 60
    return f"{mins}m {secs:.0f}s"


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------
_db_config_cache: Optional[dict] = None


def _get_db_config() -> Optional[dict]:
    """Load DB config via VaultClient. Cached after first successful call."""
    global _db_config_cache
    if _db_config_cache is not None:
        return _db_config_cache
    try:
        from utils.vault import VaultClient
        vc = VaultClient()
        _db_config_cache = vc.read_all_nested_secrets("postgres")
        return _db_config_cache
    except Exception:
        return None


def _get_conn():
    """Create a fresh psycopg2 connection with autocommit. Caller must close."""
    if psycopg2 is None:
        return None
    config = _get_db_config()
    if not config:
        return None
    try:
        conn = psycopg2.connect(
            host=config["host"],
            port=config["port"],
            dbname=config["database"],
            user=config["username"],
            password=config["password"],
            connect_timeout=5,
        )
        conn.autocommit = True
        return conn
    except Exception as e:
        import sys
        print(f"[DB CONNECT] ERROR: {e}", file=sys.stderr)
        return None


def db_ensure_table():
    """Create the history table if it doesn't exist, and migrate missing columns."""
    conn = _get_conn()
    if conn is None:
        return
    try:
        with conn.cursor() as cur:
            cur.execute(f"""
                CREATE TABLE IF NOT EXISTS {HISTORY_SCHEMA}.{HISTORY_TABLE} (
                    id              BIGSERIAL PRIMARY KEY,
                    session_id      TEXT        NOT NULL,
                    username        TEXT,
                    role            TEXT        NOT NULL,
                    content         TEXT        NOT NULL,
                    timestamp_utc   TIMESTAMPTZ NOT NULL DEFAULT now(),
                    duration_s      NUMERIC,
                    tokens_est      INTEGER,
                    model           TEXT,
                    documents       TEXT[]
                )
            """)
            # Migration: add username column first (before creating indexes on it)
            cur.execute(f"""
                ALTER TABLE {HISTORY_SCHEMA}.{HISTORY_TABLE}
                    ADD COLUMN IF NOT EXISTS username TEXT
            """)
            cur.execute(f"""
                CREATE INDEX IF NOT EXISTS idx_{HISTORY_TABLE}_session
                    ON {HISTORY_SCHEMA}.{HISTORY_TABLE} (session_id, timestamp_utc)
            """)
            cur.execute(f"""
                CREATE INDEX IF NOT EXISTS idx_{HISTORY_TABLE}_username
                    ON {HISTORY_SCHEMA}.{HISTORY_TABLE} (username)
            """)
    except Exception as e:
        import sys
        print(f"[db_ensure_table] ERROR: {e}", file=sys.stderr)
    finally:
        if conn:
            conn.close()


def db_save_message(msg: dict, session_id: str, username: str, documents: list[str]):
    """Persist a single message to postgres."""
    conn = _get_conn()
    if conn is None:
        st.toast("DB: could not connect", icon="⚠️")
        return
    try:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                INSERT INTO {HISTORY_SCHEMA}.{HISTORY_TABLE}
                    (session_id, username, role, content, timestamp_utc,
                     duration_s, tokens_est, model, documents)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    session_id,
                    username or "",
                    msg["role"],
                    msg["content"],
                    now_local(),
                    msg.get("duration"),
                    msg.get("tokens"),
                    MODEL,
                    documents or [],
                ),
            )
    except Exception as e:
        import sys
        print(f"[db_save_message] ERROR: {e}", file=sys.stderr)
        st.toast(f"DB save error: {e}", icon="⚠️")
    finally:
        conn.close()


def db_fetch_history(
    limit: int = 200,
    session_filter: Optional[str] = None,
    username_filter: Optional[str] = None,
    role_filter: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
) -> list[dict]:
    """Fetch conversation history rows for the admin view."""
    conn = _get_conn()
    if conn is None:
        return []
    try:
        clauses = []
        params: list = []
        if session_filter:
            clauses.append("session_id = %s")
            params.append(session_filter)
        if username_filter:
            clauses.append("username = %s")
            params.append(username_filter)
        if role_filter and role_filter != "All":
            clauses.append("role = %s")
            params.append(role_filter.lower())
        if date_from:
            clauses.append("timestamp_utc >= %s")
            params.append(date_from)
        if date_to:
            clauses.append("timestamp_utc <= %s")
            params.append(date_to)
        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        params.append(limit)
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                f"""
                SELECT id, session_id, username, role, content, timestamp_utc,
                       duration_s, tokens_est, model, documents
                FROM {HISTORY_SCHEMA}.{HISTORY_TABLE}
                {where}
                ORDER BY timestamp_utc DESC
                LIMIT %s
                """,
                params,
            )
            return [dict(r) for r in cur.fetchall()]
    except Exception:
        return []
    finally:
        if conn:
            conn.close()


def db_fetch_stats(
    session_filter=None, username_filter=None, date_from=None, date_to=None,
) -> dict:
    """Aggregate stats for the admin dashboard, with optional filters."""
    conn = _get_conn()
    if conn is None:
        return {}
    try:
        wheres, params = [], []
        if session_filter:
            wheres.append("session_id = %s"); params.append(session_filter)
        if username_filter:
            wheres.append("username = %s"); params.append(username_filter)
        if date_from:
            wheres.append("timestamp_utc >= %s"); params.append(date_from)
        if date_to:
            wheres.append("timestamp_utc <= %s"); params.append(date_to)
        where_clause = (" WHERE " + " AND ".join(wheres)) if wheres else ""
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(f"""
                SELECT
                    COUNT(*)                                        AS total_messages,
                    COUNT(DISTINCT session_id)                      AS total_sessions,
                    COUNT(*) FILTER (WHERE role = 'user')          AS user_messages,
                    COUNT(*) FILTER (WHERE role = 'assistant')     AS assistant_messages,
                    COALESCE(SUM(tokens_est), 0)                   AS total_tokens,
                    COALESCE(AVG(duration_s)
                        FILTER (WHERE role = 'assistant'), 0)      AS avg_duration_s,
                    COALESCE(MAX(duration_s)
                        FILTER (WHERE role = 'assistant'), 0)      AS max_duration_s,
                    MIN(timestamp_utc)                             AS first_message,
                    MAX(timestamp_utc)                             AS last_message
                FROM {HISTORY_SCHEMA}.{HISTORY_TABLE}
                {where_clause}
            """, params)
            return dict(cur.fetchone() or {})
    except Exception:
        return {}
    finally:
        if conn:
            conn.close()


def db_fetch_sessions() -> list[dict]:
    """List all distinct sessions with summary."""
    conn = _get_conn()
    if conn is None:
        return []
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(f"""
                SELECT
                    session_id,
                    MAX(username)       AS username,
                    MIN(timestamp_utc)  AS started_at,
                    MAX(timestamp_utc)  AS last_at,
                    COUNT(*)            AS message_count,
                    SUM(tokens_est)     AS total_tokens,
                    AVG(duration_s)
                        FILTER (WHERE role = 'assistant') AS avg_duration_s
                FROM {HISTORY_SCHEMA}.{HISTORY_TABLE}
                GROUP BY session_id
                ORDER BY last_at DESC
            """)
            return [dict(r) for r in cur.fetchall()]
    except Exception:
        return []
    finally:
        if conn:
            conn.close()


def db_fetch_usernames() -> list[str]:
    """Return all distinct usernames for the filter dropdown."""
    conn = _get_conn()
    if conn is None:
        return []
    try:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT DISTINCT username FROM {HISTORY_SCHEMA}.{HISTORY_TABLE}
                WHERE username IS NOT NULL AND username != ''
                ORDER BY username
            """)
            return [r[0] for r in cur.fetchall()]
    except Exception:
        return []
    finally:
        if conn:
            conn.close()


def db_clear_all() -> bool:
    """Delete all rows from the history table. Returns True on success."""
    conn = _get_conn()
    if conn is None:
        return False
    try:
        with conn.cursor() as cur:
            cur.execute(f"TRUNCATE TABLE {HISTORY_SCHEMA}.{HISTORY_TABLE}")
        return True
    except Exception:
        return False
    finally:
        if conn:
            conn.close()


def db_fetch_timeseries(
    session_filter=None, username_filter=None, date_from=None, date_to=None,
) -> list[dict]:
    """Daily aggregates for charts, with optional filters."""
    conn = _get_conn()
    if conn is None:
        return []
    try:
        wheres, params = [], []
        if session_filter:
            wheres.append("session_id = %s"); params.append(session_filter)
        if username_filter:
            wheres.append("username = %s"); params.append(username_filter)
        if date_from:
            wheres.append("timestamp_utc >= %s"); params.append(date_from)
        if date_to:
            wheres.append("timestamp_utc <= %s"); params.append(date_to)
        where_clause = (" WHERE " + " AND ".join(wheres)) if wheres else ""
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(f"""
                SELECT
                    DATE(timestamp_utc)                                          AS day,
                    COUNT(*)                                                     AS messages,
                    COUNT(DISTINCT session_id)                                   AS sessions,
                    COUNT(*) FILTER (WHERE role = 'user')                       AS user_msgs,
                    COUNT(*) FILTER (WHERE role = 'assistant')                  AS assistant_msgs,
                    COALESCE(SUM(tokens_est), 0)                                AS tokens,
                    COALESCE(AVG(duration_s) FILTER (WHERE role='assistant'), 0) AS avg_duration
                FROM {HISTORY_SCHEMA}.{HISTORY_TABLE}
                {where_clause}
                GROUP BY DATE(timestamp_utc)
                ORDER BY day
            """, params)
            return [dict(r) for r in cur.fetchall()]
    except Exception:
        return []
    finally:
        if conn:
            conn.close()


def db_fetch_user_activity(
    session_filter=None, username_filter=None, date_from=None, date_to=None,
) -> list[dict]:
    """Per-user aggregate stats, with optional filters."""
    conn = _get_conn()
    if conn is None:
        return []
    try:
        wheres, params = [], []
        if session_filter:
            wheres.append("session_id = %s"); params.append(session_filter)
        if username_filter:
            wheres.append("username = %s"); params.append(username_filter)
        if date_from:
            wheres.append("timestamp_utc >= %s"); params.append(date_from)
        if date_to:
            wheres.append("timestamp_utc <= %s"); params.append(date_to)
        where_clause = (" WHERE " + " AND ".join(wheres)) if wheres else ""
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(f"""
                SELECT
                    COALESCE(NULLIF(username, ''), '(anonymous)') AS username,
                    COUNT(*)                                      AS total_messages,
                    COUNT(DISTINCT session_id)                    AS sessions,
                    COUNT(*) FILTER (WHERE role = 'user')        AS user_msgs,
                    COUNT(*) FILTER (WHERE role = 'assistant')   AS assistant_msgs,
                    COALESCE(SUM(tokens_est), 0)                 AS tokens,
                    MIN(timestamp_utc)                            AS first_active,
                    MAX(timestamp_utc)                            AS last_active
                FROM {HISTORY_SCHEMA}.{HISTORY_TABLE}
                {where_clause}
                GROUP BY COALESCE(NULLIF(username, ''), '(anonymous)')
                ORDER BY total_messages DESC
            """, params)
            return [dict(r) for r in cur.fetchall()]
    except Exception:
        return []
    finally:
        if conn:
            conn.close()


def db_fetch_session_topics(
    limit: int = 50,
    session_filter=None, username_filter=None, date_from=None, date_to=None,
) -> list[dict]:
    """First user message per session, with optional filters."""
    conn = _get_conn()
    if conn is None:
        return []
    try:
        wheres = ["role = 'user'"]
        params: list = []
        if session_filter:
            wheres.append("session_id = %s"); params.append(session_filter)
        if username_filter:
            wheres.append("username = %s"); params.append(username_filter)
        if date_from:
            wheres.append("timestamp_utc >= %s"); params.append(date_from)
        if date_to:
            wheres.append("timestamp_utc <= %s"); params.append(date_to)
        where_clause = " WHERE " + " AND ".join(wheres)
        params.append(limit)
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(f"""
                SELECT DISTINCT ON (session_id)
                    session_id,
                    COALESCE(NULLIF(username, ''), '(anonymous)') AS username,
                    content,
                    timestamp_utc
                FROM {HISTORY_SCHEMA}.{HISTORY_TABLE}
                {where_clause}
                ORDER BY session_id, timestamp_utc ASC
                LIMIT %s
            """, params)
            return [dict(r) for r in cur.fetchall()]
    except Exception:
        return []
    finally:
        if conn:
            conn.close()


def _extract_topic_keywords(topics: list[dict], top_n: int = 20) -> list[tuple[str, int]]:
    """Simple keyword frequency from session-opening user messages."""
    STOP = {
        "the", "a", "an", "is", "are", "was", "were", "be", "been", "being",
        "have", "has", "had", "do", "does", "did", "will", "would", "shall",
        "should", "may", "might", "must", "can", "could", "i", "me", "my",
        "you", "your", "we", "our", "they", "them", "their", "he", "she",
        "it", "its", "this", "that", "these", "those", "what", "which",
        "who", "whom", "how", "when", "where", "why", "not", "no", "nor",
        "and", "or", "but", "if", "then", "so", "as", "of", "in", "on",
        "at", "to", "for", "with", "by", "from", "about", "into", "through",
        "during", "before", "after", "above", "below", "up", "down", "out",
        "off", "over", "under", "again", "further", "once", "here", "there",
        "all", "each", "every", "both", "few", "more", "most", "other",
        "some", "such", "only", "own", "same", "than", "too", "very",
        "just", "because", "also", "any", "many", "much", "like", "get",
        "got", "make", "know", "think", "want", "tell", "see", "go", "come",
        "take", "give", "use", "find", "say", "said", "let", "need", "try",
        "ask", "work", "call", "put", "keep", "still", "should", "could",
        "hi", "hello", "please", "thanks", "thank", "hey", "ok", "okay",
    }
    freq: dict[str, int] = {}
    for t in topics:
        words = re.findall(r"[a-zA-Z]{3,}", t.get("content", "").lower())
        for w in words:
            if w not in STOP:
                freq[w] = freq.get(w, 0) + 1
    return sorted(freq.items(), key=lambda x: x[1], reverse=True)[:top_n]


# ---------------------------------------------------------------------------
# Session state defaults
# ---------------------------------------------------------------------------
def _init_state():
    defaults = {
        "chat_session_id": "",
        "chat_messages": [],  # each: {role, content, timestamp, duration?, tokens?}
        "documents": {},      # name -> {content, word_count, token_count}
        "images": {},         # name -> {b64: str, mime: str}  (vision mode)
        "chat_active": False,
        "code_mode": False,
        "_generated_code": None,  # latest generated Streamlit code
        "_pending_prompt": None,  # prompt text waiting in queue
        "_pending_prompt_id": None,  # unique id for the queued prompt
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init_state()
db_ensure_table()


# ---------------------------------------------------------------------------
# Prompt-level queue — DB-based, shared across all instances
# ---------------------------------------------------------------------------
# Schema:
#   prompt_queue (id SERIAL, prompt_id TEXT UNIQUE, session_id TEXT,
#                 status TEXT [active|waiting], created_at TIMESTAMPTZ,
#                 heartbeat_at TIMESTAMPTZ)
# Concurrency: LOCK TABLE ... IN EXCLUSIVE MODE inside transactions.
# ---------------------------------------------------------------------------

def _ensure_queue_table():
    """Create the prompt_queue table if it doesn't exist."""
    conn = _get_conn()
    if conn is None:
        return
    try:
        with conn.cursor() as cur:
            cur.execute(f"""
                CREATE TABLE IF NOT EXISTS {HISTORY_SCHEMA}.{QUEUE_TABLE} (
                    id          SERIAL PRIMARY KEY,
                    prompt_id   TEXT NOT NULL UNIQUE,
                    session_id  TEXT NOT NULL,
                    status      TEXT NOT NULL DEFAULT 'waiting',
                    created_at  TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    heartbeat_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )
            """)
            cur.execute(f"""
                CREATE INDEX IF NOT EXISTS idx_{QUEUE_TABLE}_status
                ON {HISTORY_SCHEMA}.{QUEUE_TABLE} (status, created_at)
            """)
    except Exception as e:
        import sys
        print(f"[_ensure_queue_table] ERROR: {e}", file=sys.stderr)
    finally:
        conn.close()


def _queue_cleanup(cur):
    """Expire stale entries and promote next. Must be called inside a transaction."""
    tbl = f"{HISTORY_SCHEMA}.{QUEUE_TABLE}"
    # Expire stale active prompts
    cur.execute(f"""
        DELETE FROM {tbl}
        WHERE status = 'active'
          AND heartbeat_at < NOW() - INTERVAL '{PROMPT_TIMEOUT_S} seconds'
    """)
    # Expire stale waiting entries
    cur.execute(f"""
        DELETE FROM {tbl}
        WHERE status = 'waiting'
          AND created_at < NOW() - INTERVAL '{QUEUE_WAIT_TIMEOUT_S} seconds'
    """)
    # Promote next waiting → active if no active exists
    cur.execute(f"""
        SELECT COUNT(*) FROM {tbl} WHERE status = 'active'
    """)
    active_count = cur.fetchone()[0]
    if active_count == 0:
        cur.execute(f"""
            UPDATE {tbl}
            SET status = 'active', heartbeat_at = NOW()
            WHERE id = (
                SELECT id FROM {tbl}
                WHERE status = 'waiting'
                ORDER BY created_at ASC
                LIMIT 1
            )
        """)


def prompt_enqueue(session_id: str, username: str) -> tuple[str, int]:
    """Add a prompt to the queue. Returns (prompt_id, position).
    Position 0 = you are active right now. 1+ = waiting."""
    prompt_id = hashlib.sha256(
        f"{session_id}:{time.time()}:{os.getpid()}".encode()
    ).hexdigest()[:16]

    conn = _get_conn()
    if conn is None:
        return prompt_id, 0  # fallback: no DB → proceed without queue

    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            # Lock the table to prevent race conditions
            cur.execute(f"LOCK TABLE {HISTORY_SCHEMA}.{QUEUE_TABLE} IN EXCLUSIVE MODE")
            _queue_cleanup(cur)

            # Check if there's an active prompt
            cur.execute(f"""
                SELECT COUNT(*) FROM {HISTORY_SCHEMA}.{QUEUE_TABLE}
                WHERE status = 'active'
            """)
            has_active = cur.fetchone()[0] > 0

            if not has_active:
                # Go straight to active
                cur.execute(f"""
                    INSERT INTO {HISTORY_SCHEMA}.{QUEUE_TABLE}
                        (prompt_id, session_id, status, created_at, heartbeat_at)
                    VALUES (%s, %s, 'active', NOW(), NOW())
                """, (prompt_id, session_id))
                conn.commit()
                return prompt_id, 0
            else:
                # Add to waiting queue
                cur.execute(f"""
                    INSERT INTO {HISTORY_SCHEMA}.{QUEUE_TABLE}
                        (prompt_id, session_id, status, created_at, heartbeat_at)
                    VALUES (%s, %s, 'waiting', NOW(), NOW())
                """, (prompt_id, session_id))
                # Get position (count of waiting entries ahead of this one)
                cur.execute(f"""
                    SELECT COUNT(*) FROM {HISTORY_SCHEMA}.{QUEUE_TABLE}
                    WHERE status = 'waiting' AND prompt_id != %s
                      AND created_at <= (
                          SELECT created_at FROM {HISTORY_SCHEMA}.{QUEUE_TABLE}
                          WHERE prompt_id = %s
                      )
                """, (prompt_id, prompt_id))
                pos = cur.fetchone()[0] + 1  # +1 because the active one is pos 0
                conn.commit()
                return prompt_id, pos
    except Exception as e:
        import sys
        print(f"[prompt_enqueue] ERROR: {e}", file=sys.stderr)
        try:
            conn.rollback()
        except Exception:
            pass
        return prompt_id, 0  # fallback
    finally:
        conn.autocommit = True
        conn.close()


def prompt_position(prompt_id: str) -> int:
    """Get queue position for a prompt_id.
    0 = active, 1+ = waiting, -1 = not found (expired or done).
    Read-only — no exclusive lock."""
    conn = _get_conn()
    if conn is None:
        return -1
    try:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT status FROM {HISTORY_SCHEMA}.{QUEUE_TABLE}
                WHERE prompt_id = %s
            """, (prompt_id,))
            row = cur.fetchone()
            if row is None:
                return -1
            if row[0] == 'active':
                return 0
            # Waiting — get position
            cur.execute(f"""
                SELECT COUNT(*) FROM {HISTORY_SCHEMA}.{QUEUE_TABLE}
                WHERE status = 'waiting'
                  AND created_at <= (
                      SELECT created_at FROM {HISTORY_SCHEMA}.{QUEUE_TABLE}
                      WHERE prompt_id = %s
                  )
            """, (prompt_id,))
            return cur.fetchone()[0]
    except Exception as e:
        import sys
        print(f"[prompt_position] ERROR: {e}", file=sys.stderr)
        return -1
    finally:
        conn.close()


def prompt_release(prompt_id: str):
    """Release the active slot after prompt processing is done."""
    conn = _get_conn()
    if conn is None:
        return
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            cur.execute(f"LOCK TABLE {HISTORY_SCHEMA}.{QUEUE_TABLE} IN EXCLUSIVE MODE")
            cur.execute(f"""
                DELETE FROM {HISTORY_SCHEMA}.{QUEUE_TABLE}
                WHERE prompt_id = %s
            """, (prompt_id,))
            _queue_cleanup(cur)  # promotes next in line
            conn.commit()
    except Exception as e:
        import sys
        print(f"[prompt_release] ERROR: {e}", file=sys.stderr)
        try:
            conn.rollback()
        except Exception:
            pass
    finally:
        conn.autocommit = True
        conn.close()


def prompt_cancel(prompt_id: str):
    """Remove a prompt from the queue (if user cancels while waiting)."""
    conn = _get_conn()
    if conn is None:
        return
    try:
        with conn.cursor() as cur:
            cur.execute(f"""
                DELETE FROM {HISTORY_SCHEMA}.{QUEUE_TABLE}
                WHERE prompt_id = %s
            """, (prompt_id,))
    except Exception as e:
        import sys
        print(f"[prompt_cancel] ERROR: {e}", file=sys.stderr)
    finally:
        conn.close()


def prompt_heartbeat(prompt_id: str):
    """Refresh the heartbeat timestamp for the active prompt (keep-alive)."""
    conn = _get_conn()
    if conn is None:
        return
    try:
        with conn.cursor() as cur:
            cur.execute(f"""
                UPDATE {HISTORY_SCHEMA}.{QUEUE_TABLE}
                SET heartbeat_at = NOW()
                WHERE prompt_id = %s AND status = 'active'
            """, (prompt_id,))
    except Exception as e:
        import sys
        print(f"[prompt_heartbeat] ERROR: {e}", file=sys.stderr)
    finally:
        conn.close()


def prompt_queue_status() -> dict:
    """Return the full queue state (for display). Anonymous — no usernames.
    Read-only — no exclusive lock, just reads current state."""
    conn = _get_conn()
    if conn is None:
        return {"active": None, "queue": []}
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Lightweight cleanup: delete expired rows (autocommit handles it)
            cur.execute(f"""
                DELETE FROM {HISTORY_SCHEMA}.{QUEUE_TABLE}
                WHERE (status = 'active'
                       AND heartbeat_at < NOW() - INTERVAL '{PROMPT_TIMEOUT_S} seconds')
                   OR (status = 'waiting'
                       AND created_at < NOW() - INTERVAL '{QUEUE_WAIT_TIMEOUT_S} seconds')
            """)

            cur.execute(f"""
                SELECT prompt_id, session_id, status, created_at, heartbeat_at
                FROM {HISTORY_SCHEMA}.{QUEUE_TABLE}
                WHERE status = 'active'
                LIMIT 1
            """)
            active_row = cur.fetchone()

            cur.execute(f"""
                SELECT prompt_id, session_id, status, created_at, heartbeat_at
                FROM {HISTORY_SCHEMA}.{QUEUE_TABLE}
                WHERE status = 'waiting'
                ORDER BY created_at ASC
            """)
            waiting_rows = cur.fetchall()

        active = None
        if active_row:
            active = {
                "prompt_id": active_row["prompt_id"],
                "session_id": active_row["session_id"],
                "started_at": active_row["heartbeat_at"].isoformat()
                    if active_row["heartbeat_at"] else active_row["created_at"].isoformat(),
            }
        queue = []
        for w in waiting_rows:
            queue.append({
                "prompt_id": w["prompt_id"],
                "session_id": w["session_id"],
                "queued_at": w["created_at"].isoformat(),
            })
        return {"active": active, "queue": queue}
    except Exception as e:
        import sys
        print(f"[prompt_queue_status] ERROR: {e}", file=sys.stderr)
        return {"active": None, "queue": []}
    finally:
        conn.close()


_ensure_queue_table()


# ---------------------------------------------------------------------------
# Ollama helpers
# ---------------------------------------------------------------------------
def ollama_is_running() -> bool:
    try:
        r = requests.get(f"{OLLAMA_URL}/api/tags", timeout=3)
        return r.status_code == 200
    except Exception:
        return False


def _ollama_model_exists(model_name: str) -> bool:
    """Check if a specific model is available on the Ollama instance."""
    try:
        r = requests.get(f"{OLLAMA_URL}/api/tags", timeout=3)
        if r.status_code != 200:
            return False
        models = [m["name"] for m in r.json().get("models", [])]
        # Match exact or without tag (e.g. "qwen2.5-vl:7b" matches "qwen2.5-vl:7b")
        return model_name in models or any(m.startswith(model_name.split(":")[0] + ":") for m in models if model_name.split(":")[0] in m)
    except Exception:
        return False


def chat_stream(messages: list[dict], model: str | None = None):
    payload = {"model": model or MODEL, "messages": messages, "stream": True}
    with requests.post(f"{OLLAMA_URL}/api/chat", json=payload, stream=True, timeout=120) as r:
        r.raise_for_status()
        for line in r.iter_lines():
            if line:
                chunk = json.loads(line)
                token = chunk.get("message", {}).get("content", "")
                if token:
                    yield token
                if chunk.get("done"):
                    break


# ---------------------------------------------------------------------------
# .doc binary text extraction helpers
# ---------------------------------------------------------------------------
import struct


def _extract_doc_text_olefile(raw: bytes) -> str:
    """
    Parse a .doc file using olefile by reading the FIB (File Information Block)
    and the piece table from the table stream to reconstruct the document text.
    """
    ole = olefile.OleFileIO(BytesIO(raw))
    try:
        # Read the WordDocument stream
        word_stream = ole.openstream("WordDocument")
        word_data = word_stream.read()

        if len(word_data) < 1024:
            return ""

        # FIB: read key offsets
        # Bytes 10-11: flags (bit 9 = fWhichTblStm, tells us 0Table vs 1Table)
        fib_flags = struct.unpack_from("<H", word_data, 0x000A)[0]
        use_1table = bool(fib_flags & 0x0200)
        table_name = "1Table" if use_1table else "0Table"

        if not ole.exists(table_name):
            return ""

        table_stream = ole.openstream(table_name)
        table_data = table_stream.read()

        # FIB: CLX (complex part) offset and size at 0x01A2 and 0x01A6
        clx_offset = struct.unpack_from("<I", word_data, 0x01A2)[0]
        clx_size = struct.unpack_from("<I", word_data, 0x01A6)[0]

        if clx_offset == 0 or clx_size == 0:
            return ""

        clx_data = table_data[clx_offset: clx_offset + clx_size]

        # Walk the CLX to find the Pcdt (piece table descriptor)
        pos = 0
        piece_table = None
        while pos < len(clx_data):
            clxt = clx_data[pos]
            if clxt == 0x02:  # Pcdt
                pcdt_size = struct.unpack_from("<I", clx_data, pos + 1)[0]
                piece_table = clx_data[pos + 5: pos + 5 + pcdt_size]
                break
            elif clxt == 0x01:  # Grpprl — skip
                grpprl_size = struct.unpack_from("<H", clx_data, pos + 1)[0]
                pos += 3 + grpprl_size
            else:
                break

        if piece_table is None:
            return ""

        # Parse piece table: array of CPs followed by array of PCDs
        # Number of pieces = (size - 4) / (4 + 8) ... each CP is 4 bytes, each PCD is 8 bytes
        # Actually: n+1 CPs (4 bytes each) + n PCDs (8 bytes each) where n = piece count
        # Total = (n+1)*4 + n*8 = 4 + 12n => n = (len - 4) / 12
        pt_len = len(piece_table)
        n_pieces = (pt_len - 4) // 12
        if n_pieces <= 0:
            return ""

        # Read character positions (n+1 of them)
        cps = []
        for i in range(n_pieces + 1):
            cp = struct.unpack_from("<I", piece_table, i * 4)[0]
            cps.append(cp)

        # Read piece descriptors (starting after CPs)
        pcd_offset = (n_pieces + 1) * 4
        text_parts = []
        for i in range(n_pieces):
            pcd_start = pcd_offset + i * 8
            if pcd_start + 8 > pt_len:
                break
            # PCD: 2 bytes flags, 4 bytes fc (file char position), 2 bytes prm
            fc_raw = struct.unpack_from("<I", piece_table, pcd_start + 2)[0]
            is_ansi = bool(fc_raw & 0x40000000)
            fc = fc_raw & 0x3FFFFFFF
            char_count = cps[i + 1] - cps[i]

            if is_ansi:
                # ANSI: 1 byte per character, fc is divided by 2
                start = fc // 2
                end = start + char_count
                if end <= len(word_data):
                    text_parts.append(
                        word_data[start:end].decode("cp1252", errors="replace")
                    )
            else:
                # Unicode: 2 bytes per character
                start = fc
                end = start + char_count * 2
                if end <= len(word_data):
                    text_parts.append(
                        word_data[start:end].decode("utf-16-le", errors="replace")
                    )

        full_text = "".join(text_parts)
        # Clean up Word control characters
        # \r = paragraph mark, \x07 = cell/row mark, \x0c = page break
        full_text = full_text.replace("\r", "\n")
        full_text = full_text.replace("\x07", "\t")
        full_text = full_text.replace("\x0c", "\n\n")
        full_text = full_text.replace("\x01", "")  # field begin
        full_text = full_text.replace("\x13", "")  # field separator
        full_text = full_text.replace("\x14", "")  # field separator
        full_text = full_text.replace("\x15", "")  # field end
        # Collapse excessive newlines
        full_text = re.sub(r'\n{3,}', '\n\n', full_text)
        return full_text.strip()
    except Exception:
        return ""
    finally:
        ole.close()


def _extract_doc_text_bruteforce(raw: bytes) -> str:
    """
    Last-resort: try multiple decodings of the entire binary and extract
    the longest readable text runs.
    """
    best = ""
    for encoding in ("utf-16-le", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding, errors="replace")
        except Exception:
            continue
        # Find runs of printable chars (including common punctuation and unicode)
        runs = re.findall(r'[\w \t.,;:!?\'\"()\-/\n]{10,}', text, re.UNICODE)
        # Keep only runs that look like prose (have spaces)
        prose = [r.strip() for r in runs if " " in r and len(r.strip()) > 20]
        candidate = "\n".join(prose)
        if len(candidate) > len(best):
            best = candidate
    return best


# ---------------------------------------------------------------------------
# Document parsing
# ---------------------------------------------------------------------------
def _extract_excel_text(raw: bytes, name: str) -> str:
    """Extract text content from .xlsx / .xls files, one section per sheet."""
    # .xlsx via openpyxl (already a dependency)
    if name.endswith(".xlsx"):
        if Workbook is None:
            return "[ERROR] openpyxl is not installed. Run: pip install openpyxl"
        from openpyxl import load_workbook
        try:
            wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        except Exception as e:
            return f"[ERROR] Cannot read .xlsx file: {e}"
        sections = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows.append("\t".join(cells))
            if rows:
                # Build a readable header + rows representation
                header = f"=== Sheet: {sheet_name} ({len(rows)} rows) ==="
                sections.append(f"{header}\n" + "\n".join(rows))
        wb.close()
        if sections:
            return "\n\n".join(sections)
        return "[WARNING] Excel file has no data."

    # .xls via xlrd (optional) or fallback message
    if name.endswith(".xls"):
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=raw)
            sections = []
            for sheet in wb.sheets():
                rows = []
                for rx in range(sheet.nrows):
                    cells = [str(sheet.cell_value(rx, cx)) for cx in range(sheet.ncols)]
                    if any(c.strip() for c in cells):
                        rows.append("\t".join(cells))
                if rows:
                    header = f"=== Sheet: {sheet.name} ({len(rows)} rows) ==="
                    sections.append(f"{header}\n" + "\n".join(rows))
            if sections:
                return "\n\n".join(sections)
            return "[WARNING] Excel file has no data."
        except ImportError:
            return "[ERROR] xlrd is not installed for .xls support. Run: pip install xlrd"
        except Exception as e:
            return f"[ERROR] Cannot read .xls file: {e}"

    return "[ERROR] Unsupported Excel format."


def extract_text(file) -> str:
    name = file.name.lower()
    file.seek(0)
    raw = file.read()

    if name.endswith(".txt") or name.endswith(".md"):
        return raw.decode("utf-8", errors="replace")

    if name.endswith(".pdf"):
        pdf_bytes = BytesIO(raw)
        if pdfplumber is not None:
            pages = []
            with pdfplumber.open(pdf_bytes) as pdf:
                for i, page in enumerate(pdf.pages, 1):
                    text = page.extract_text() or ""
                    if text.strip():
                        pages.append(f"--- Page {i} ---\n{text}")
            if pages:
                return "\n\n".join(pages)
            return "[WARNING] No text could be extracted from this PDF (it may be scanned/image-based)."
        if PdfReader is not None:
            reader = PdfReader(pdf_bytes)
            pages = []
            for i, page in enumerate(reader.pages, 1):
                text = page.extract_text() or ""
                if text.strip():
                    pages.append(f"--- Page {i} ---\n{text}")
            if pages:
                return "\n\n".join(pages)
            return "[WARNING] No text could be extracted from this PDF (it may be scanned/image-based)."
        return "[ERROR] No PDF library installed. Run: pip install pdfplumber"

    if name.endswith(".xlsx") or name.endswith(".xls"):
        return _extract_excel_text(raw, name)

    if name.endswith(".docx"):
        if DocxDocument is None:
            return "[ERROR] python-docx is not installed. Run: pip install python-docx"
        doc = DocxDocument(BytesIO(raw))
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())

    if name.endswith(".doc"):
        # Strategy 1: antiword CLI (most reliable)
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".doc", delete=False) as tmp:
                tmp.write(raw)
                tmp_path = tmp.name
        except Exception:
            pass
        if tmp_path:
            try:
                result = subprocess.run(
                    ["antiword", "-m", "UTF-8", tmp_path],
                    capture_output=True, text=True, timeout=30,
                )
                if result.returncode == 0 and result.stdout.strip():
                    os.unlink(tmp_path)
                    return result.stdout.strip()
            except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
                pass
            # Strategy 2: libreoffice headless
            try:
                subprocess.run(
                    ["libreoffice", "--headless", "--convert-to", "txt:Text",
                     "--outdir", "/tmp", tmp_path],
                    capture_output=True, text=True, timeout=30,
                )
                txt_out = "/tmp/" + os.path.basename(tmp_path).replace(".doc", ".txt")
                if os.path.exists(txt_out):
                    with open(txt_out, "r", errors="replace") as f:
                        content = f.read()
                    os.unlink(txt_out)
                    if content.strip():
                        os.unlink(tmp_path)
                        return content.strip()
            except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
                pass
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

        # Strategy 3: pure-Python .doc parser via olefile + piece table
        if olefile is not None:
            try:
                text = _extract_doc_text_olefile(raw)
                if text:
                    return text
            except Exception:
                pass

        # Strategy 4: brute-force binary text extraction (last resort)
        text = _extract_doc_text_bruteforce(raw)
        if text:
            return text

        return "[ERROR] Cannot extract text from this .doc file. Try converting it to .docx."

    return f"[Unsupported file type: {name}]"


def add_document(file):
    """Extract text, compute stats, store in session state. Returns True if new."""
    if file.name in st.session_state.documents:
        return False
    content = extract_text(file)
    wc = count_words(content)
    tc = estimate_tokens(content)
    st.session_state.documents[file.name] = {
        "content": content,
        "word_count": wc,
        "token_count": tc,
    }
    return True


# ---------------------------------------------------------------------------
# Build system prompt with document context
# ---------------------------------------------------------------------------
def build_system_prompt() -> str:
    username = st.session_state.get("username", "")
    title = st.session_state.get("title", "")
    teams = st.session_state.get("teams", [])
    roles = st.session_state.get("roles", [])
    parts = [
        "You are a helpful, professional assistant. Answer the user's questions clearly and accurately.",
    ]
    user_details = []
    if username:
        user_details.append(f"Name: {username}")
    if title:
        user_details.append(f"Title: {title}")
    if teams:
        teams_str = ", ".join(teams) if isinstance(teams, list) else str(teams)
        user_details.append(f"Teams: {teams_str}")
    if roles:
        roles_str = ", ".join(roles) if isinstance(roles, list) else str(roles)
        user_details.append(f"Roles: {roles_str}")
    if user_details:
        parts.append(
            "USER CONTEXT (internal — use to subtly tailor your responses):\n"
            + "\n".join(f"- {d}" for d in user_details)
            + "\n\nIMPORTANT: Do NOT mention these details explicitly in your responses "
            "unless the user specifically asks about them. Use this context silently "
            "to adjust the depth, terminology, and relevance of your answers to match "
            "the user's expertise and organizational context."
        )
    if st.session_state.documents:
        parts.append(
            "\nThe user has uploaded the following documents. "
            "Use their content to answer questions when relevant.\n"
        )
        for doc_name, doc_info in st.session_state.documents.items():
            text = doc_info["content"]
            truncated = text[:80_000]
            suffix = "... [truncated]" if len(text) > 80_000 else ""
            parts.append(f"### Document: {doc_name}\n```\n{truncated}{suffix}\n```\n")
    return "\n".join(parts)


def build_code_system_prompt() -> str:
    """System prompt for Streamlit page-builder mode."""
    username = st.session_state.get("username", "")
    title = st.session_state.get("title", "")
    user_ctx = ""
    if username or title:
        user_ctx = f"\nThe user is {username}" + (f", a {title}" if title else "") + ".\n"

    doc_ctx = ""
    if st.session_state.documents:
        doc_ctx = "\nThe user has uploaded these documents whose data you can reference:\n"
        for doc_name, doc_info in st.session_state.documents.items():
            text = doc_info["content"][:20_000]
            suffix = "... [truncated]" if len(doc_info["content"]) > 20_000 else ""
            doc_ctx += f"### {doc_name}\n```\n{text}{suffix}\n```\n"

    current_code = st.session_state.get("_generated_code") or ""
    code_ctx = ""
    if current_code:
        code_ctx = f"""

CURRENT PAGE CODE (already running — the user wants to modify THIS page):
```python
{current_code}
```

EDITING RULES:
- The user is asking you to edit the EXISTING page above.
- You MUST return the COMPLETE updated code with the requested changes applied.
- Do NOT generate a new page from scratch — modify the existing code.
- Keep all parts the user did NOT mention unchanged.
- Fix only what the user asks to fix. Do not refactor or restyle untouched sections.
"""

    return f"""You are an expert Streamlit developer. Your ONLY job is to generate complete, runnable Streamlit Python code.
{user_ctx}{doc_ctx}{code_ctx}
RULES — follow these strictly:
1. Respond ONLY with a single Python code block (```python ... ```). No explanations before or after.
2. The code will be executed inside an existing Streamlit page via exec(). Do NOT call st.set_page_config().
3. The code must be self-contained — it will be exec'd as-is.
4. ALWAYS return the FULL page code, even when making small edits. The output replaces the current page entirely.
5. Import all needed packages at the top of the code block: pandas, altair, datetime, random, math, etc.
6. Do NOT use st.set_page_config, st.sidebar, st.cache_resource, or st.cache_data in the generated code.
7. If the user asks about data from their uploaded documents, parse and visualize that data.
8. Always include realistic sample/mock data if the user's request requires data you don't have.

STREAMLIT BEST PRACTICES — follow these to produce clean, error-free, professional pages:

Layout & Structure:
- Use st.columns() for side-by-side layouts. Always unpack correctly: col1, col2 = st.columns(2). Never nest columns inside columns (Streamlit does not support it).
- Use st.tabs() for multi-section pages. Unpack as: tab1, tab2 = st.tabs(["Tab A", "Tab B"]), then use `with tab1:`.
- Use st.container() to group related elements. Use st.expander() for collapsible detail sections.
- Place st.metric() inside columns for KPI card rows. Use the `delta` parameter for trend indicators.
- Add vertical spacing with st.markdown("") or st.divider() — never st.write("").

Data & DataFrames:
- Always construct DataFrames with explicit column names: pd.DataFrame({{"Col": [...]}}).
- Use st.dataframe(df, use_container_width=True, hide_index=True) for clean tables.
- For editable tables use st.data_editor(). For static display prefer st.dataframe().
- When showing metrics from a DataFrame, use .iloc[0] or .values[0] to extract scalars — never pass a Series to st.metric.

Charts & Visualization:
- For simple charts: st.bar_chart(), st.line_chart(), st.area_chart() accept DataFrames with index as x-axis.
- For advanced charts: use altair (import altair as alt). Build with alt.Chart(df).mark_*().encode(). Always call .properties(height=N) to set chart height.
- Set chart colors explicitly — do not rely on defaults. Use the `color` parameter in st.bar_chart/st.line_chart or alt.Color in altair.
- For horizontal bars use `horizontal=True` parameter in st.bar_chart().

Styling & UX:
- Inject custom CSS via st.markdown('<style>...</style>', unsafe_allow_html=True) at the TOP of the code.
- Use card-style layouts: white background, subtle border (1px solid #e0e0e0), border-radius 12px, padding 1rem, box-shadow for elevation.
- Use a consistent color palette. Good defaults: primary #4a90d9, success #2d8a4e, warning #e6a817, danger #c0392b, background #f5f2eb, card #ffffff.
- Add section headers with st.markdown("### Section Title") — not st.header() which is too large.
- Use st.caption() for subtle secondary text. Use st.markdown() with unsafe_allow_html for fine-grained HTML.

Forms & Input:
- Group related inputs with st.form() and st.form_submit_button() to prevent constant reruns.
- Use appropriate widgets: st.selectbox for <10 options, st.multiselect for multi-choice, st.slider for numeric ranges, st.date_input for dates.
- Always provide sensible default values for all inputs.

Common Pitfalls to AVOID:
- NEVER use st.columns() inside st.columns() — it raises an error.
- NEVER pass a list or Series where a scalar is expected (st.metric value, st.progress value).
- NEVER call st.set_page_config() — the host page already called it.
- NEVER use st.sidebar — this is rendered inside a page section.
- NEVER use st.cache_resource or st.cache_data — the code is exec'd fresh each time.
- NEVER use st.button() to toggle state — buttons only return True on the click rerun. Use st.toggle() or st.checkbox() for persistent state.
- NEVER use f-strings inside st.markdown HTML with curly braces in CSS — escape them or use .format(). Example: use `st.markdown('<style>.card {{padding: 1rem}}</style>', unsafe_allow_html=True)` (double braces).
- When using st.tabs(), always match the number of unpacked variables to the number of tab labels.
- Always handle empty DataFrames gracefully — check `if not df.empty:` before charting or computing stats.
- When building altair charts, always specify .encode(x=..., y=...) explicitly — never rely on altair to guess.
- String concatenation in st.markdown: use + operator or .format(), not f-strings when the template contains literal CSS braces.

Execution Safety:
- The generated page must ONLY provide a UI — forms, inputs, buttons, visualizations.
- NEVER execute actions automatically on page load (no API calls, no file writes, no database queries, no network requests, no subprocess calls that run on import).
- All actions (data processing, calculations, API calls, file operations) must be gated behind a user-triggered event: a st.form_submit_button(), st.button(), or similar interactive widget.
- Computed results, charts, and tables should only appear AFTER the user clicks a button or submits a form.
- Use st.session_state to store results so they persist across reruns, but never auto-populate on first load.
- NEVER use subprocess, os.system, requests, urllib, or any I/O module that executes on import.
- Mock or simulate any external data — do not make real HTTP requests or database connections."""


def _extract_code_block(text: str) -> Optional[str]:
    """Extract the first Python code block from LLM output."""
    # Try ```python ... ``` first
    pattern = r"```python\s*\n(.*?)```"
    m = re.search(pattern, text, re.DOTALL)
    if m:
        return m.group(1).strip()
    # Fallback: any ``` block
    pattern = r"```\s*\n(.*?)```"
    m = re.search(pattern, text, re.DOTALL)
    if m:
        return m.group(1).strip()
    return None


def _execute_generated_code(code: str):
    """Safely execute generated Streamlit code in a container."""
    st.markdown(
        '<div class="code-output-panel">'
        '<div class="code-output-header"><span>Generated Page</span></div>'
        '<div class="code-output-body">',
        unsafe_allow_html=True,
    )
    try:
        # Build a controlled exec environment
        exec_globals = {"__builtins__": __builtins__}
        exec_globals["st"] = st
        exec_globals["pd"] = None
        exec_globals["np"] = None
        try:
            import pandas as pd
            exec_globals["pd"] = pd
        except ImportError:
            pass
        try:
            import numpy as np
            exec_globals["np"] = np
        except ImportError:
            pass
        try:
            import altair as alt
            exec_globals["alt"] = alt
        except ImportError:
            pass
        exec(code, exec_globals)
    except Exception as e:
        st.error(f"Code execution error: {e}")
        import traceback
        st.code(traceback.format_exc(), language="python")
    st.markdown('</div></div>', unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Export conversation
# ---------------------------------------------------------------------------
def export_conversation_md() -> str:
    """Generate markdown export from current session state. Always reads fresh."""
    messages = list(st.session_state.get("chat_messages", []))
    documents = dict(st.session_state.get("documents", {}))

    lines = [
        f"# Chat Export — {now_local().strftime('%Y-%m-%d %H:%M')}",
        f"**Model:** {MODEL}",
        f"**Documents:** {', '.join(documents.keys()) or 'None'}",
        f"**Messages:** {len(messages)}",
        "",
        "---",
        "",
    ]
    for i, msg in enumerate(messages, 1):
        role = "You" if msg["role"] == "user" else "Assistant"
        ts = msg.get("timestamp", "")
        dur = msg.get("duration")
        tokens = msg.get("tokens")
        meta_parts = []
        if ts:
            meta_parts.append(ts)
        if dur is not None:
            meta_parts.append(format_duration(dur))
        if tokens is not None:
            meta_parts.append(f"~{format_number(tokens)} tokens")
        meta = f" ({' · '.join(meta_parts)})" if meta_parts else ""
        lines.append(f"### {role}{meta}\n\n{msg['content']}\n")
    return "\n".join(lines)


def _parse_markdown_tables(text: str) -> list[tuple[str, list[list[str]]]]:
    """
    Find markdown tables in text. Returns list of (context_label, rows)
    where rows is a list of lists (header + data rows).
    """
    tables = []
    lines = text.split("\n")
    i = 0
    table_idx = 0
    while i < len(lines):
        line = lines[i].strip()
        # A markdown table row starts and ends with |
        if line.startswith("|") and line.endswith("|"):
            # Collect all consecutive table lines
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith("|") and lines[i].strip().endswith("|"):
                table_lines.append(lines[i].strip())
                i += 1
            if len(table_lines) >= 2:  # At least header + separator
                rows = []
                for tl in table_lines:
                    cells = [c.strip() for c in tl.strip("|").split("|")]
                    # Skip separator rows (---|---|---)
                    if all(re.match(r'^[-:]+$', c) for c in cells):
                        continue
                    rows.append(cells)
                if rows:
                    table_idx += 1
                    tables.append((f"Table {table_idx}", rows))
        else:
            i += 1
    return tables


def export_conversation_xlsx() -> Optional[bytes]:
    """
    Generate Excel export. Conversation goes on the main sheet.
    Any markdown tables found in assistant responses get their own tabs.
    Returns None if openpyxl is not installed or no messages.
    """
    if Workbook is None:
        return None
    messages = list(st.session_state.get("chat_messages", []))
    if not messages:
        return None

    wb = Workbook()

    # -- Sheet 1: Conversation --
    ws = wb.active
    ws.title = "Conversation"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    border = Border(
        bottom=Side(style="thin", color="D5D0C4"),
        right=Side(style="thin", color="D5D0C4"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = ["#", "Role", "Time", "Message", "Duration", "~Tokens"]
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    all_tables = []
    for i, msg in enumerate(messages, 1):
        role = "You" if msg["role"] == "user" else "Assistant"
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=role)
        ws.cell(row=row, column=3, value=msg.get("timestamp", ""))
        content_cell = ws.cell(row=row, column=4, value=msg["content"])
        content_cell.alignment = wrap
        dur = msg.get("duration")
        ws.cell(row=row, column=5, value=format_duration(dur) if dur is not None else "")
        tokens = msg.get("tokens")
        ws.cell(row=row, column=6, value=tokens if tokens is not None else "")

        for cell_ref in ws[row]:
            cell_ref.border = border

        # Collect tables from assistant messages
        if msg["role"] == "assistant":
            found = _parse_markdown_tables(msg["content"])
            for label, rows in found:
                all_tables.append((f"Msg{i} {label}", rows))

    # -- Additional sheets for tables --
    for sheet_name, rows in all_tables:
        # Sheet names max 31 chars
        safe_name = sheet_name[:31]
        ts = wb.create_sheet(title=safe_name)
        for r_idx, row_data in enumerate(rows, 1):
            for c_idx, val in enumerate(row_data, 1):
                cell = ts.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                    cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border
        # Auto-width columns
        for col_cells in ts.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ts.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Message content renderer — collapsible text & code snippets
# ---------------------------------------------------------------------------
_CODE_BLOCK_RE = re.compile(r'```(\w*)\n(.*?)```', re.DOTALL)


def _render_collapsible_text(text: str, key: str):
    """Render text with collapse if it exceeds preview thresholds."""
    lines = text.split('\n')
    is_long = len(lines) > MAX_PREVIEW_LINES or len(text) > MAX_PREVIEW_CHARS

    if not is_long:
        st.markdown(text)
        return

    # Build truncated preview
    preview = '\n'.join(lines[:MAX_PREVIEW_LINES])
    if len(preview) > MAX_PREVIEW_CHARS:
        preview = preview[:MAX_PREVIEW_CHARS].rsplit(' ', 1)[0]

    st.markdown(preview + " ...\n\n---")
    with st.expander("Show full response", expanded=False):
        st.markdown(text)


def render_message_content(content: str, msg_key: str):
    """Render a message with collapsible long text and hidden code snippets."""
    # Extract code blocks
    code_blocks: list[tuple[str, str]] = []
    for match in _CODE_BLOCK_RE.finditer(content):
        lang = match.group(1) or ''
        code = match.group(2).rstrip()
        code_blocks.append((lang, code))

    # Text with code blocks removed
    text_only = _CODE_BLOCK_RE.sub('', content).strip()
    text_only = re.sub(r'\n{3,}', '\n\n', text_only)

    # Render text portion (collapsible if long)
    if text_only:
        _render_collapsible_text(text_only, msg_key)

    # Render each code block as a collapsible snippet with copy support
    for i, (lang, code) in enumerate(code_blocks):
        lines = code.count('\n') + 1
        lang_label = lang if lang else "code"
        with st.expander(f"`{lang_label}` — {lines} lines", expanded=False):
            st.code(code, language=lang or None)


# ---------------------------------------------------------------------------
# Message metadata renderer
# ---------------------------------------------------------------------------
def render_msg_meta(msg: dict):
    """Render subtle timestamp / duration / token info below a message."""
    parts = []
    ts = msg.get("timestamp")
    if ts:
        parts.append(f"<span>{ts}</span>")
    dur = msg.get("duration")
    if dur is not None:
        parts.append(f"<span>{format_duration(dur)}</span>")
    tokens = msg.get("tokens")
    if tokens is not None:
        parts.append(f"<span>~{format_number(tokens)} tokens</span>")
    if parts:
        st.markdown(f'<div class="msg-meta">{"".join(parts)}</div>', unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# LOBBY VIEW
# ---------------------------------------------------------------------------
def render_lobby():
    st.markdown("## Document Chat")
    st.caption("Chat with your documents using local Ollama models")

    online = ollama_is_running()
    q_status = prompt_queue_status()
    active_prompt = q_status.get("active")
    waiting_count = len(q_status.get("queue", []))

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("#### Chatbot Status")

        if online:
            st.markdown(
                f'<span class="status-badge status-active">Ollama Online</span> '
                f'<span class="toolbar-meta">{MODEL}</span>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<span class="status-badge status-offline">Ollama Offline</span>',
                unsafe_allow_html=True,
            )

        st.write("")

        # Show queue status info
        if active_prompt:
            st.markdown(
                '<span class="status-badge status-busy">Processing</span> '
                '<span class="toolbar-meta">A prompt is being processed</span>',
                unsafe_allow_html=True,
            )
            if waiting_count:
                st.caption(f"{waiting_count} prompt{'s' if waiting_count != 1 else ''} waiting in queue")
        else:
            st.markdown(
                '<span class="status-badge status-active">Available</span>',
                unsafe_allow_html=True,
            )
            #st.caption("No prompts running — your next message will process immediately.")

        st.write("")
        if st.button("Start Chat", use_container_width=True, key="start_chat", type="primary"):
            st.session_state.chat_session_id = hashlib.sha256(
                f"{time.time()}".encode()
            ).hexdigest()[:16]
            st.session_state.chat_active = True
            st.rerun()

        st.markdown('</div>', unsafe_allow_html=True)



# ---------------------------------------------------------------------------
# TOOLBAR
# ---------------------------------------------------------------------------
def render_toolbar():
    online = ollama_is_running()
    doc_count = len(st.session_state.documents)
    is_admin = "admin" in st.session_state.get("user_roles", {})
    total_tokens = sum(d["token_count"] for d in st.session_state.documents.values())
    code_mode = st.session_state.code_mode

    # --- Top bar: brand + action buttons ---
    cols = []
    if is_admin:
        cols = st.columns([3, 1.2, 1.2, 1.2, 1.2, 1.2])
    else:
        cols = st.columns([4, 1.2, 1.2, 1.2, 1.2])

    col_idx = 0

    # Brand + queue status
    with cols[col_idx]:
        status_cls = "status-active" if online else "status-offline"
        status_txt = "Connected" if online else "Offline"
        brand_parts = (
            f'<div class="toolbar-row">'
            f'<span class="toolbar-brand">Document Chat</span> '
            f'<span class="status-badge {status_cls}">{status_txt}</span>'
        )
        if is_admin and code_mode:
            brand_parts += ' <span class="mode-pill mode-code">Page Builder</span>'
        if ENABLE_VISION and st.session_state.images:
            brand_parts += f' <span class="status-badge status-active">Vision · {VISION_MODEL}</span>'
        # Queue indicator
        q_status = prompt_queue_status()
        q_active = q_status.get("active")
        q_waiting = len(q_status.get("queue", []))
        if q_active:
            brand_parts += f' <span class="status-badge status-busy">Queue: {q_waiting + 1}</span>'
        brand_parts += '</div>'
        st.markdown(brand_parts, unsafe_allow_html=True)
    col_idx += 1

    # Page Builder toggle (admin only)
    if is_admin:
        with cols[col_idx]:
            if st.toggle("Page Builder", value=code_mode, key="code_mode_toggle",
                          help="Switch to Streamlit page generation mode"):
                if not st.session_state.code_mode:
                    st.session_state.code_mode = True
                    st.rerun()
            else:
                if st.session_state.code_mode:
                    st.session_state.code_mode = False
                    st.rerun()
        col_idx += 1

    # Clear conversation
    with cols[col_idx]:
        if st.button("Clear Chat", use_container_width=True, key="chat_clear"):
            st.session_state.chat_messages = []
            st.session_state.pop("_export_md", None)
            st.session_state.pop("_export_xlsx", None)
            st.rerun()
    col_idx += 1

    # Export
    with cols[col_idx]:
        if st.button("Export", use_container_width=True, key="chat_prepare_export"):
            st.session_state["_export_md"] = export_conversation_md()
            st.session_state["_export_xlsx"] = export_conversation_xlsx()
            st.rerun()
    col_idx += 1

    # End session
    with cols[col_idx]:
        if st.button("End Session", use_container_width=True, key="chat_leave", type="primary"):
            # Cancel any pending queued prompt
            pid = st.session_state.get("_pending_prompt_id")
            if pid:
                prompt_cancel(pid)
            st.session_state.chat_active = False
            st.session_state.chat_messages = []
            st.session_state.documents = {}
            st.session_state.images = {}
            st.session_state._pending_prompt = None
            st.session_state._pending_prompt_id = None
            st.session_state.pop("_export_md", None)
            st.session_state.pop("_export_xlsx", None)
            st.rerun()

    # --- Download row (appears after Export is clicked) ---
    md_data = st.session_state.get("_export_md")
    xlsx_data = st.session_state.get("_export_xlsx")
    if md_data or xlsx_data:
        dl_cols = st.columns([3, 1.5, 1.5])
        with dl_cols[0]:
            st.caption(f"{len(st.session_state.chat_messages)} messages ready to download")
        if md_data:
            with dl_cols[1]:
                st.download_button(
                    "Markdown",
                    data=md_data,
                    file_name=f"chat_export_{now_local().strftime('%Y%m%d_%H%M%S')}.md",
                    mime="text/markdown",
                    use_container_width=True,
                    key="chat_export_md",
                )
        if xlsx_data:
            with dl_cols[2]:
                st.download_button(
                    "Excel",
                    data=xlsx_data,
                    file_name=f"chat_export_{now_local().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="chat_export_xlsx",
                )

    # --- Upload row ---
    _upload_types = ["txt", "md", "pdf", "doc", "docx", "xlsx", "xls"]
    if ENABLE_VISION:
        _upload_types += ["jpg", "jpeg", "png", "bmp"]
    uploaded = st.file_uploader(
        "Upload documents" + (" & images" if ENABLE_VISION else ""),
        type=_upload_types,
        accept_multiple_files=True,
        key="chat_doc_uploader",
        label_visibility="collapsed",
    )
    if uploaded:
        new_added = False
        for f in uploaded:
            ext = Path(f.name).suffix.lower()
            if ext in (".jpg", ".jpeg", ".png", ".bmp") and ENABLE_VISION:
                # Store image as base64 for vision API
                if f.name not in st.session_state.images:
                    if not hasattr(st.session_state, "_vision_model_warned") and not _ollama_model_exists(VISION_MODEL):
                        st.toast(f"Vision model **{VISION_MODEL}** not found. Pull it with: `ollama pull {VISION_MODEL}`", icon="⚠️")
                        st.session_state._vision_model_warned = True
                    raw = f.read()
                    mime = {"jpg": "image/jpeg", "jpeg": "image/jpeg",
                            "png": "image/png", "bmp": "image/bmp"}.get(ext.lstrip("."), "image/jpeg")
                    st.session_state.images[f.name] = {
                        "b64": base64.b64encode(raw).decode("utf-8"),
                        "mime": mime,
                        "size": len(raw),
                    }
                    new_added = True
            else:
                if add_document(f):
                    new_added = True
        if new_added:
            st.rerun()

    # --- Document & image chips bar with remove buttons ---
    has_docs = bool(st.session_state.documents)
    has_imgs = bool(st.session_state.images)
    if has_docs or has_imgs:
        chips_html = '<div class="doc-bar">'
        if has_docs:
            chips_html += '<span class="doc-bar-label">Docs:</span>'
            for doc_name, doc_info in st.session_state.documents.items():
                ext = Path(doc_name).suffix.lower()
                icon = {".txt": "📄", ".md": "📝", ".pdf": "📕", ".doc": "📙", ".docx": "📘", ".xlsx": "📊", ".xls": "📊"}.get(ext, "📎")
                chips_html += (
                    f'<span class="doc-chip">{icon} {doc_name}'
                    f'<span style="color:var(--text-muted);font-size:0.7rem;margin-left:4px;">'
                    f'{format_number(doc_info["word_count"])}w · ~{format_number(doc_info["token_count"])}tok</span></span>'
                )
        if has_imgs:
            chips_html += '<span class="doc-bar-label" style="margin-left:8px;">Images:</span>'
            for img_name, img_info in st.session_state.images.items():
                size_kb = img_info["size"] / 1024
                size_str = f"{size_kb:.0f}KB" if size_kb < 1024 else f"{size_kb/1024:.1f}MB"
                chips_html += (
                    f'<span class="doc-chip">🖼️ {img_name}'
                    f'<span style="color:var(--text-muted);font-size:0.7rem;margin-left:4px;">{size_str}</span></span>'
                )
        chips_html += "</div>"
        st.markdown(chips_html, unsafe_allow_html=True)

        # Remove buttons row
        all_items = (
            [(n, "doc") for n in st.session_state.documents]
            + [(n, "img") for n in st.session_state.images]
        )
        rm_cols = st.columns(len(all_items))
        for i, (name, kind) in enumerate(all_items):
            with rm_cols[i]:
                if st.button(f"✕ {name[:20]}", key=f"rm_{kind}_{name}",
                             use_container_width=True, help=f"Remove {name}"):
                    if kind == "doc":
                        del st.session_state.documents[name]
                    else:
                        del st.session_state.images[name]
                    st.rerun()

    st.markdown('<hr class="divider" style="margin:0.25rem 0 0.5rem;">', unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# CHAT AREA
# ---------------------------------------------------------------------------
def _generate_greeting():
    """Stream a personalized greeting from the LLM based on user identity."""
    username = st.session_state.get("username", "")
    title = st.session_state.get("title", "")
    doc_count = len(st.session_state.documents)

    system = build_system_prompt()
    user_parts = []
    if username:
        user_parts.append(f"my name is {username}")
    if title:
        user_parts.append(f"I work as a {title}")
    if doc_count:
        doc_names = ", ".join(list(st.session_state.documents.keys())[:5])
        user_parts.append(f"I have uploaded {doc_count} document(s): {doc_names}")

    greeting_prompt = (
        "Greet me briefly and warmly in 1-2 sentences. "
        "Mention my name if you know it, acknowledge my role if you know it, "
        "and mention my uploaded documents if any. "
        "Then ask how you can help. Keep it concise and professional."
    )
    if user_parts:
        greeting_prompt = f"Context: {', '.join(user_parts)}. " + greeting_prompt

    api_messages = [
        {"role": "system", "content": system},
        {"role": "user", "content": greeting_prompt},
    ]

    with st.chat_message("assistant"):
        placeholder = st.empty()
        meta_placeholder = st.empty()
        full_response = ""
        t_start = time.time()
        try:
            for token in chat_stream(api_messages):
                full_response += token
                placeholder.markdown(full_response + "▌")
            placeholder.markdown(full_response)
        except Exception:
            full_response = f"Hello{' ' + username if username else ''}! How can I help you today?"
            placeholder.markdown(full_response)

        duration = time.time() - t_start
        resp_tokens = estimate_tokens(full_response)
        resp_ts = now_local().strftime("%H:%M")
        assistant_msg = {
            "role": "assistant",
            "content": full_response,
            "timestamp": resp_ts,
            "duration": duration,
            "tokens": resp_tokens,
        }
        meta_placeholder.markdown(
            f'<div class="msg-meta">'
            f'<span>{resp_ts}</span>'
            f'<span>{format_duration(duration)}</span>'
            f'<span>~{format_number(resp_tokens)} tokens</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

    st.session_state.chat_messages.append(assistant_msg)
    db_save_message(assistant_msg, st.session_state.chat_session_id,
                    st.session_state.get("username", ""),
                    list(st.session_state.documents.keys()))


def _render_chat_conversation():
    """Render the conversation messages and empty state."""
    doc_count = len(st.session_state.documents)
    username = st.session_state.get("username", "")
    code_mode = st.session_state.code_mode

    if not st.session_state.chat_messages:
        if code_mode:
            st.markdown(
                '<div class="empty-state">'
                '<h2>Page Builder Mode</h2>'
                '<p>Describe the Streamlit page you want and I\'ll generate it live. '
                'Try: "Build a sales dashboard with KPI cards and charts"</p>'
                '</div>',
                unsafe_allow_html=True,
            )
        else:
            display_name = username if username else "there"
            greeting = f"Hello, {display_name}!"
            hint = (
                "Upload documents via the Documents button above, or just start a conversation."
                if not doc_count
                else "Your documents are loaded and ready. Start asking questions below."
            )
            st.markdown(
                f'<div class="empty-state"><h2>{greeting}</h2><p>{hint}</p></div>',
                unsafe_allow_html=True,
            )

    for idx, msg in enumerate(st.session_state.chat_messages):
        if msg["role"] == "user":
            with st.chat_message("user"):
                render_message_content(msg["content"], f"usr_{idx}")
                # Show attached image thumbnails
                attached = msg.get("_images", [])
                if attached:
                    img_cols = st.columns(min(len(attached), 4))
                    for j, img_data in enumerate(attached):
                        with img_cols[j % len(img_cols)]:
                            st.image(
                                base64.b64decode(img_data["b64"]),
                                caption=img_data.get("name", ""),
                                use_container_width=True,
                            )
                render_msg_meta(msg)
        else:
            if code_mode and _extract_code_block(msg["content"]):
                with st.chat_message("assistant"):
                    st.markdown("Page updated.")
                    render_msg_meta(msg)
            else:
                with st.chat_message("assistant"):
                    render_message_content(msg["content"], f"msg_{idx}")
                    render_msg_meta(msg)



def _render_generated_page():
    """Render the generated Streamlit page with source code viewer."""
    code = st.session_state.get("_generated_code")
    if not code:
        st.markdown(
            '<div class="empty-state">'
            '<h2>No page generated yet</h2>'
            '<p>Switch to the Chat tab and describe the page you want to build. '
            'The generated page will appear here.</p>'
            '</div>',
            unsafe_allow_html=True,
        )
        return
    with st.expander("View source code", expanded=False):
        st.code(code, language="python")
    _execute_generated_code(code)


def _render_queue_wait(position: int, prompt_id: str):
    """Display a styled queue waiting card with position and auto-refresh."""
    ordinal = (
        f"{position}st" if position == 1 else
        f"{position}nd" if position == 2 else
        f"{position}rd" if position == 3 else
        f"{position}th"
    )

    st.markdown(
        f'<div class="queue-card">'
        f'<div class="queue-position">#{position}</div>'
        f'<div class="queue-label">You are {ordinal} in line</div>'
        f'<div class="queue-hint">'
        f'Another prompt is currently being processed. '
        f'Your message will be sent automatically when it\'s your turn.'
        f'</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    qc1, qc2, qc3 = st.columns([1, 1, 1])
    with qc2:
        if st.button("Cancel", use_container_width=True, key="cancel_queue"):
            prompt_cancel(prompt_id)
            # Remove the pending user message from chat
            if st.session_state.chat_messages and st.session_state.chat_messages[-1]["role"] == "user":
                st.session_state.chat_messages.pop()
            st.session_state._pending_prompt = None
            st.session_state._pending_prompt_id = None
            st.rerun()

    # Auto-refresh every 2 seconds to check queue
    time.sleep(2)
    st.rerun()


def render_chat():
    doc_count = len(st.session_state.documents)
    code_mode = st.session_state.code_mode
    has_page = bool(st.session_state.get("_generated_code"))

    # When in code mode, show Chat and Generated Page as separate tabs
    if code_mode:
        tab_labels = ["Chat", "Generated Page"]
        tab_chat, tab_page = st.tabs(tab_labels)
        with tab_chat:
            _render_chat_conversation()
        with tab_page:
            _render_generated_page()
    else:
        _render_chat_conversation()

    # Disclaimer — always visible above chat input
    st.markdown(
        '<div style="text-align:center;padding:0.4rem 0;font-size:0.73rem;'
        'color:#e67e22;font-weight:500;letter-spacing:0.01em;">'
        'This chat is monitored — conversations are logged for quality assurance and performance tracking.'
        '</div>',
        unsafe_allow_html=True,
    )

    # Chat input (always visible below tabs)
    has_images = ENABLE_VISION and bool(st.session_state.images)
    if code_mode:
        placeholder_text = "Describe the page you want to build..."
    elif has_images and doc_count:
        placeholder_text = "Ask about your documents and images..."
    elif has_images:
        placeholder_text = "Ask about your images..."
    elif doc_count:
        placeholder_text = "Ask about your documents..."
    else:
        placeholder_text = "Ask me anything..."

    # ------------------------------------------------------------------
    # Saved prompts — team + role quick actions
    # ------------------------------------------------------------------
    user_teams = st.session_state.get("teams", [])
    user_roles = st.session_state.get("user_roles", {})
    available_prompts: list[tuple[str, str]] = []  # (label, prompt)
    seen_labels: set[str] = set()
    for team in user_teams:
        for sp in TEAM_PROMPTS.get(team, []):
            if sp["label"] not in seen_labels:
                available_prompts.append((sp["label"], sp["prompt"]))
                seen_labels.add(sp["label"])
    for role in user_roles:
        for sp in ROLE_PROMPTS.get(role, []):
            if sp["label"] not in seen_labels:
                available_prompts.append((sp["label"], sp["prompt"]))
                seen_labels.add(sp["label"])

    if available_prompts and not code_mode:
        st.markdown(
            '<div class="saved-prompts-bar">'
            '<div class="saved-prompts-label">Quick Prompts</div>'
            '</div>',
            unsafe_allow_html=True,
        )
        sp_cols = st.columns(min(len(available_prompts), 4))
        for i, (label, sp_text) in enumerate(available_prompts):
            with sp_cols[i % len(sp_cols)]:
                if st.button(label, key=f"sp_{i}", use_container_width=True):
                    st.session_state["_saved_prompt_fire"] = sp_text
                    st.rerun()

    # Pick up a fired saved prompt (set on previous run)
    fired_prompt = st.session_state.pop("_saved_prompt_fire", None)

    # ------------------------------------------------------------------
    # Handle pending prompt waiting in queue
    # ------------------------------------------------------------------
    pending_pid = st.session_state.get("_pending_prompt_id")
    if pending_pid:
        pos = prompt_position(pending_pid)
        if pos > 0:
            # Still waiting — show queue card and auto-refresh
            _render_queue_wait(pos, pending_pid)
            return
        elif pos == -1:
            # Expired / cancelled — clear pending state
            st.session_state._pending_prompt = None
            st.session_state._pending_prompt_id = None
            pending_pid = None
        # pos == 0 → it's our turn — fall through to process below

    # ------------------------------------------------------------------
    # Chat input (typed or saved prompt)
    # ------------------------------------------------------------------
    prompt = st.chat_input(placeholder_text, key="chat_input") or fired_prompt

    if prompt:
        if not ollama_is_running():
            st.error("Ollama is not reachable. Check the connection.")
            return

        # Save user message immediately (visible in history)
        user_ts = now_local().strftime("%H:%M")
        user_tokens = estimate_tokens(prompt)
        user_msg = {
            "role": "user",
            "content": prompt,
            "timestamp": user_ts,
            "tokens": user_tokens,
        }
        # Attach images snapshot to message for re-rendering
        if ENABLE_VISION and st.session_state.images:
            user_msg["_images"] = [
                {"name": n, "b64": info["b64"], "mime": info["mime"]}
                for n, info in st.session_state.images.items()
            ]
        st.session_state.chat_messages.append(user_msg)
        db_save_message(user_msg, st.session_state.chat_session_id,
                        st.session_state.get("username", ""),
                        list(st.session_state.documents.keys()))

        # Render user message immediately (conversation loop already ran above)
        with st.chat_message("user"):
            st.markdown(prompt)
            if user_msg.get("_images"):
                img_cols = st.columns(min(len(user_msg["_images"]), 4))
                for j, img_data in enumerate(user_msg["_images"]):
                    with img_cols[j % len(img_cols)]:
                        st.image(
                            base64.b64decode(img_data["b64"]),
                            caption=img_data.get("name", ""),
                            use_container_width=True,
                        )
            render_msg_meta(user_msg)

        # Enqueue the prompt
        username = st.session_state.get("username", "")
        pid, pos = prompt_enqueue(st.session_state.chat_session_id, username)

        if pos > 0:
            # Not our turn yet — save and rerun to show queue card
            st.session_state._pending_prompt = prompt
            st.session_state._pending_prompt_id = pid
            st.rerun()
            return
        else:
            # Our turn — save pid so we can release later, then process
            st.session_state._pending_prompt = prompt
            st.session_state._pending_prompt_id = pid

    # ------------------------------------------------------------------
    # Process the prompt (either fresh pos==0 or resumed from queue)
    # ------------------------------------------------------------------
    pending_prompt = st.session_state.get("_pending_prompt")
    pending_prompt_id = st.session_state.get("_pending_prompt_id")

    if pending_prompt and pending_prompt_id:
        # Build API messages
        if code_mode:
            system_content = build_code_system_prompt()
        else:
            system_content = build_system_prompt()
        system_msg = {"role": "system", "content": system_content}

        # Determine if we should use vision model
        use_vision = ENABLE_VISION and bool(st.session_state.images)
        if use_vision and not _ollama_model_exists(VISION_MODEL):
            st.warning(
                f"Vision model **{VISION_MODEL}** is not available on Ollama. "
                f"Images will be ignored. Pull it with: `ollama pull {VISION_MODEL}`"
            )
            use_vision = False
        stream_model = VISION_MODEL if use_vision else None  # None = default MODEL

        # Build conversation messages, attaching images to the latest user message
        api_messages = [system_msg]
        for i, m in enumerate(st.session_state.chat_messages):
            msg_entry = {"role": m["role"], "content": m["content"]}
            # Attach images to the last user message only
            if use_vision and m["role"] == "user" and i == len(st.session_state.chat_messages) - 1:
                msg_entry["images"] = [
                    img["b64"] for img in st.session_state.images.values()
                ]
            api_messages.append(msg_entry)

        # Show queue bar while processing (clearable)
        processing_bar = st.empty()
        processing_bar.markdown(
            '<div class="queue-bar">Processing your prompt...</div>',
            unsafe_allow_html=True,
        )

        with st.chat_message("assistant"):
            placeholder = st.empty()
            meta_placeholder = st.empty()
            full_response = ""
            t_start = time.time()
            try:
                for token in chat_stream(api_messages, model=stream_model):
                    full_response += token
                    placeholder.markdown(full_response + "▌")
                    # Keep-alive heartbeat every ~10 tokens
                    if len(full_response) % 40 == 0:
                        prompt_heartbeat(pending_prompt_id)
                placeholder.markdown(full_response)
            except requests.exceptions.ConnectionError:
                full_response = "Connection to Ollama lost. Please check that it is running."
                placeholder.error(full_response)
            except Exception as e:
                full_response = f"Error: {e}"
                placeholder.error(full_response)

            duration = time.time() - t_start
            resp_tokens = estimate_tokens(full_response)
            resp_ts = now_local().strftime("%H:%M")
            assistant_msg = {
                "role": "assistant",
                "content": full_response,
                "timestamp": resp_ts,
                "duration": duration,
                "tokens": resp_tokens,
            }
            meta_placeholder.markdown(
                f'<div class="msg-meta">'
                f'<span>{resp_ts}</span>'
                f'<span>{format_duration(duration)}</span>'
                f'<span>~{format_number(resp_tokens)} tokens</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

        # Clear the processing bar
        processing_bar.empty()

        st.session_state.chat_messages.append(assistant_msg)
        db_save_message(assistant_msg, st.session_state.chat_session_id,
                        st.session_state.get("username", ""),
                        list(st.session_state.documents.keys()))

        # Release the queue slot
        prompt_release(pending_prompt_id)
        st.session_state._pending_prompt = None
        st.session_state._pending_prompt_id = None

        # In code mode, extract code and update the single page
        if code_mode:
            extracted_code = _extract_code_block(full_response)
            if extracted_code:
                st.session_state["_generated_code"] = extracted_code
                st.rerun()
            else:
                st.warning("No code block found in the response. Try rephrasing your request.")


# ---------------------------------------------------------------------------
# Admin view
# ---------------------------------------------------------------------------
def render_admin():
    """Full conversation history dashboard — visible to admins only, in lobby only."""
    st.markdown('<div class="history-section">', unsafe_allow_html=True)

    # -- Section header --
    hc1, hc2 = st.columns([6, 1])
    with hc1:
        st.markdown(
            '<div class="history-section-title">'
            '<h3>Conversation History</h3>'
            '<span class="admin-badge">Admin</span>'
            '</div>',
            unsafe_allow_html=True,
        )
    with hc2:
        if st.button("Refresh", use_container_width=True, key="admin_refresh"):
            st.rerun()

    if psycopg2 is None:
        st.warning("psycopg2 is not installed. Run: `pip install psycopg2-binary`")
        st.markdown('</div>', unsafe_allow_html=True)
        return

    _test_conn = _get_conn()
    if _test_conn is None:
        st.warning("Could not connect to the database. Check your Vault / DB configuration.")
        st.markdown('</div>', unsafe_allow_html=True)
        return
    _test_conn.close()

    # ==========================================================
    # LIVE QUEUE MONITOR
    # ==========================================================
    q_data = prompt_queue_status()
    q_active = q_data.get("active")
    q_waiting = q_data.get("queue", [])

    dot_cls = "qm-dot" if q_active else "qm-dot qm-dot-idle"
    total_in_queue = (1 if q_active else 0) + len(q_waiting)
    title_extra = f" — {total_in_queue} prompt{'s' if total_in_queue != 1 else ''}" if total_in_queue else " — idle"

    st.markdown(
        f'<div class="queue-monitor">'
        f'<div class="queue-monitor-title">'
        f'<span class="{dot_cls}"></span> Prompt Queue{title_extra}'
        f'</div>',
        unsafe_allow_html=True,
    )

    if not q_active and not q_waiting:
        st.markdown(
            '<div style="font-size:0.8rem;color:var(--text-muted);padding:0.25rem 0;">'
            'No prompts in the queue — the model is available.'
            '</div>',
            unsafe_allow_html=True,
        )
    else:
        if q_active:
            try:
                started = datetime.fromisoformat(q_active["started_at"])
                elapsed = int((now_local() - to_local(started)).total_seconds())
                time_str = f"started {elapsed}s ago"
            except Exception:
                time_str = ""
            st.markdown(
                f'<div class="qm-entry qm-entry-active">'
                f'<span class="qm-pos qm-pos-active">NOW</span>'
                f'<span class="qm-user">Active prompt</span>'
                f'<span class="qm-detail">{time_str}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

        for i, entry in enumerate(q_waiting):
            try:
                queued = datetime.fromisoformat(entry["queued_at"])
                wait = int((now_local() - to_local(queued)).total_seconds())
                wait_str = f"waiting {wait}s"
            except Exception:
                wait_str = ""
            st.markdown(
                f'<div class="qm-entry qm-entry-waiting">'
                f'<span class="qm-pos qm-pos-waiting">#{i + 1}</span>'
                f'<span class="qm-user">Queued prompt</span>'
                f'<span class="qm-detail">{wait_str}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

    st.markdown('</div>', unsafe_allow_html=True)

    # ==========================================================
    # FILTERS — up top so they control everything below
    # ==========================================================
    sessions  = db_fetch_sessions()
    usernames = db_fetch_usernames()

    session_options  = ["All sessions"] + [s["session_id"] for s in sessions]
    username_options = ["All users"] + usernames

    st.markdown(
        '<div class="filter-bar">'
        '<div class="filter-bar-label">Filters</div>',
        unsafe_allow_html=True,
    )
    fc1, fc2, fc3, fc4 = st.columns([2.2, 1.6, 1.2, 2.4])
    with fc1:
        sel_session = st.selectbox(
            "Session",
            session_options,
            format_func=lambda s: s if s == "All sessions" else f"{s[:18]}…",
            key="admin_session_filter",
        )
    with fc2:
        sel_username = st.selectbox("User", username_options, key="admin_username_filter")
    with fc3:
        sel_role = st.selectbox("Role", ["All", "User", "Assistant"], key="admin_role_filter")
    with fc4:
        date_range_options = ["All Time", "Today", "1W", "1M", "1Y", "5Y"]
        sel_date_range = st.selectbox(
            "Period", date_range_options,
            key="admin_date_range",
        )
    st.markdown('</div>', unsafe_allow_html=True)

    # -- Resolve date range to from/to --
    now = now_local()
    date_range_map = {
        "Today": timedelta(days=1),
        "1W":    timedelta(weeks=1),
        "1M":    timedelta(days=30),
        "1Y":    timedelta(days=365),
        "5Y":    timedelta(days=365 * 5),
    }
    f_date_from = None
    f_date_to = None
    if sel_date_range and sel_date_range in date_range_map:
        f_date_from = now - date_range_map[sel_date_range]

    # -- Resolve common filter kwargs --
    f_session  = None if sel_session == "All sessions" else sel_session
    f_username = None if sel_username == "All users" else sel_username
    fkw = dict(session_filter=f_session, username_filter=f_username,
               date_from=f_date_from, date_to=f_date_to)

    # Show active filter hint
    active_filters = []
    if f_session:
        active_filters.append(f"session {f_session[:14]}…")
    if f_username:
        active_filters.append(f"user: {f_username}")
    if sel_date_range and sel_date_range != "All Time":
        active_filters.append(f"period: {sel_date_range}")
    if active_filters:
        st.markdown(
            f'<div style="font-size:0.75rem;color:var(--accent);margin-bottom:0.5rem;">'
            f'Showing filtered results: {" · ".join(active_filters)}'
            f'</div>',
            unsafe_allow_html=True,
        )

    # ==========================================================
    # STATS GRID (6 cards) — filtered
    # ==========================================================
    stats = db_fetch_stats(**fkw)
    avg_dur   = float(stats.get("avg_duration_s") or 0)
    max_dur   = float(stats.get("max_duration_s") or 0)
    first_msg = stats.get("first_message")
    last_msg  = stats.get("last_message")

    if stats:
        cs = st.columns(6)
        cards = [
            ("Sessions",       format_number(int(stats.get("total_sessions", 0))),
             f"First: {to_local(first_msg).strftime('%b %d') if first_msg else '—'}"),
            ("Total Messages", format_number(int(stats.get("total_messages", 0))),
             f"Last: {to_local(last_msg).strftime('%b %d') if last_msg else '—'}"),
            ("User",           format_number(int(stats.get("user_messages", 0))),
             "messages sent"),
            ("Assistant",      format_number(int(stats.get("assistant_messages", 0))),
             "responses given"),
            ("Avg Response",   format_duration(avg_dur),
             f"Max: {format_duration(max_dur)}"),
            ("Total Tokens",   format_number(int(stats.get("total_tokens", 0))),
             "estimated"),
        ]
        for col, (label, value, sub) in zip(cs, cards):
            with col:
                st.markdown(
                    f'<div class="stat-card">'
                    f'<div class="stat-value">{value}</div>'
                    f'<div class="stat-label">{label}</div>'
                    f'<div class="stat-sub">{sub}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

    # ==========================================================
    # CHARTS — filtered
    # ==========================================================
    ts_rows = db_fetch_timeseries(**fkw)
    if ts_rows:
        try:
            import pandas as pd
            df = pd.DataFrame(ts_rows)
            df["day"] = pd.to_datetime(df["day"])
            df = df.sort_values("day")
            df["avg_duration"] = df["avg_duration"].astype(float).round(2)
            df["tokens"] = df["tokens"].astype(int)

            st.markdown(
                '<div class="chart-section-divider"><span>Usage Over Time</span></div>',
                unsafe_allow_html=True,
            )

            ch1, ch2 = st.columns(2)

            with ch1:
                st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                st.markdown('<div class="chart-card-title">Messages per Day</div>', unsafe_allow_html=True)
                chart_df = df.set_index("day")[["user_msgs", "assistant_msgs"]].rename(
                    columns={"user_msgs": "User", "assistant_msgs": "Assistant"}
                )
                st.bar_chart(chart_df, color=["#4a90d9", "#2d8a4e"], height=200)
                st.markdown('</div>', unsafe_allow_html=True)

            with ch2:
                st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                st.markdown('<div class="chart-card-title">Avg Response Time (s)</div>', unsafe_allow_html=True)
                st.line_chart(
                    df.set_index("day")[["avg_duration"]].rename(columns={"avg_duration": "Avg (s)"}),
                    color=["#4a90d9"],
                    height=200,
                )
                st.markdown('</div>', unsafe_allow_html=True)

            ch3, ch4 = st.columns(2)

            with ch3:
                st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                st.markdown('<div class="chart-card-title">Sessions per Day</div>', unsafe_allow_html=True)
                st.bar_chart(
                    df.set_index("day")[["sessions"]].rename(columns={"sessions": "Sessions"}),
                    color=["#4a90d9"],
                    height=180,
                )
                st.markdown('</div>', unsafe_allow_html=True)

            with ch4:
                st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                st.markdown('<div class="chart-card-title">Tokens per Day (estimated)</div>', unsafe_allow_html=True)
                st.area_chart(
                    df.set_index("day")[["tokens"]].rename(columns={"tokens": "Tokens"}),
                    color=["#4a90d9"],
                    height=180,
                )
                st.markdown('</div>', unsafe_allow_html=True)

        except Exception:
            pass

    # ==========================================================
    # USER ACTIVITY & TOPICS — filtered
    # ==========================================================
    user_rows = db_fetch_user_activity(**fkw)
    session_topics = db_fetch_session_topics(**fkw)

    if user_rows or session_topics:
        try:
            import pandas as pd

            if user_rows:
                st.markdown(
                    '<div class="chart-section-divider"><span>User Activity</span></div>',
                    unsafe_allow_html=True,
                )

                ua1, ua2 = st.columns(2)

                with ua1:
                    st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                    st.markdown(
                        '<div class="chart-card-title">Messages by User</div>',
                        unsafe_allow_html=True,
                    )
                    udf = pd.DataFrame(user_rows)
                    chart_udf = udf.set_index("username")[["user_msgs", "assistant_msgs"]].rename(
                        columns={"user_msgs": "Sent", "assistant_msgs": "Received"}
                    )
                    st.bar_chart(chart_udf, color=["#4a90d9", "#2d8a4e"], horizontal=True, height=max(160, len(udf) * 38))
                    st.markdown('</div>', unsafe_allow_html=True)

                with ua2:
                    st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                    st.markdown(
                        '<div class="chart-card-title">Tokens by User</div>',
                        unsafe_allow_html=True,
                    )
                    st.bar_chart(
                        udf.set_index("username")[["tokens"]].rename(columns={"tokens": "Tokens"}),
                        color=["#4a90d9"],
                        horizontal=True,
                        height=max(160, len(udf) * 38),
                    )
                    st.markdown('</div>', unsafe_allow_html=True)

                # User detail table
                st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                st.markdown(
                    '<div class="chart-card-title">User Details</div>',
                    unsafe_allow_html=True,
                )
                table_data = []
                for u in user_rows:
                    last_ts = u["last_active"]
                    table_data.append({
                        "User": u["username"],
                        "Sessions": int(u["sessions"]),
                        "Messages": int(u["total_messages"]),
                        "Sent": int(u["user_msgs"]),
                        "Received": int(u["assistant_msgs"]),
                        "Tokens": format_number(int(u["tokens"])),
                        "First Active": to_local(u["first_active"]).strftime("%Y-%m-%d %H:%M") if u["first_active"] else "—",
                        "Last Active": to_local(last_ts).strftime("%Y-%m-%d %H:%M") if last_ts else "—",
                    })
                st.dataframe(
                    pd.DataFrame(table_data),
                    use_container_width=True,
                    hide_index=True,
                )
                st.markdown('</div>', unsafe_allow_html=True)

            # Topic analysis
            if session_topics:
                st.markdown(
                    '<div class="chart-section-divider"><span>Topics &amp; Keywords</span></div>',
                    unsafe_allow_html=True,
                )

                tp1, tp2 = st.columns(2)

                with tp1:
                    keywords = _extract_topic_keywords(session_topics, top_n=15)
                    if keywords:
                        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                        st.markdown(
                            '<div class="chart-card-title">Top Keywords (from user prompts)</div>',
                            unsafe_allow_html=True,
                        )
                        # Sort by count descending so the chart renders highest first
                        kw_df = pd.DataFrame(keywords, columns=["Keyword", "Count"])
                        kw_df = kw_df.sort_values("Count", ascending=True)  # ascending for horizontal bar (top = highest)
                        kw_df = kw_df.set_index("Keyword")
                        st.bar_chart(kw_df, color=["#4a90d9"], horizontal=True, height=max(180, len(keywords) * 28))
                        st.markdown('</div>', unsafe_allow_html=True)

                with tp2:
                    st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                    st.markdown(
                        '<div class="chart-card-title">Session Openers (first question per session)</div>',
                        unsafe_allow_html=True,
                    )
                    for t in sorted(session_topics, key=lambda x: x["timestamp_utc"] or datetime.min, reverse=True)[:15]:
                        preview = t["content"][:120].replace("\n", " ").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                        if len(t["content"]) > 120:
                            preview += " …"
                        ts_str = to_local(t["timestamp_utc"]).strftime("%b %d, %H:%M") if t["timestamp_utc"] else ""
                        st.markdown(
                            f'<div style="padding:0.45rem 0;border-bottom:1px solid var(--border-light);">'
                            f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:2px;">'
                            f'<span style="font-size:0.76rem;font-weight:600;color:var(--text-primary);">{t["username"]}</span>'
                            f'<span style="font-size:0.68rem;color:var(--text-muted);">{ts_str}</span>'
                            f'</div>'
                            f'<div style="font-size:0.8rem;color:var(--text-secondary);line-height:1.45;">{preview}</div>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
                    st.markdown('</div>', unsafe_allow_html=True)
        except Exception:
            pass

    # ==========================================================
    # MESSAGE HISTORY — filtered
    # ==========================================================
    st.markdown(
        '<div class="chart-section-divider"><span>Message History</span></div>',
        unsafe_allow_html=True,
    )

    rows = db_fetch_history(
        limit=500,
        session_filter=f_session,
        username_filter=f_username,
        role_filter=sel_role,
        date_from=f_date_from,
        date_to=f_date_to,
    )

    # -- Danger zone (admin only, behind flag) --
    if ENABLE_DANGER_ZONE:
        with st.expander("Danger Zone", expanded=False):
            st.warning("This will permanently delete **all** conversation history from the database.")
            confirm = st.checkbox("I understand, delete all history", key="admin_clear_confirm")
            if st.button(
                "Clear All History",
                disabled=not confirm,
                use_container_width=True,
                type="primary",
                key="admin_clear_all",
            ):
                if db_clear_all():
                    st.success("All history deleted.")
                    st.rerun()
                else:
                    st.error("Failed to delete history. Check DB connection.")

    # -- Result count --
    if not rows:
        st.markdown(
            '<div style="text-align:center;padding:2rem;color:var(--text-muted);font-size:0.85rem;">'
            'No messages match the selected filters.'
            '</div>',
            unsafe_allow_html=True,
        )
        st.markdown('</div>', unsafe_allow_html=True)
        return

    n_sessions = len({r["session_id"] for r in rows})
    st.markdown(
        f'<div class="stat-meta-row">'
        f'<span><strong>{len(rows)}</strong> messages</span>'
        f'<span><strong>{n_sessions}</strong> session{"s" if n_sessions != 1 else ""}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # -- Tabs --
    tab_sessions, tab_flat = st.tabs(["By Session", "Flat Log"])

    # ---- By Session tab ----
    with tab_sessions:
        grouped: dict[str, list[dict]] = {}
        for r in rows:
            grouped.setdefault(r["session_id"], []).append(r)

        for sid, msgs in grouped.items():
            n_msgs    = len(msgs)
            first_ts  = msgs[-1]["timestamp_utc"]
            last_ts   = msgs[0]["timestamp_utc"]
            total_tok = sum(m["tokens_est"] or 0 for m in msgs)
            uname     = msgs[0].get("username") or "—"
            docs_list = msgs[0].get("documents") or []

            date_str = to_local(first_ts).strftime("%Y-%m-%d") if first_ts else "—"
            time_range = ""
            if first_ts and last_ts:
                time_range = f"{to_local(first_ts).strftime('%H:%M')} – {to_local(last_ts).strftime('%H:%M')}"

            label = f"{uname}  ·  {sid[:14]}…  ·  {n_msgs} messages  ·  {date_str}"
            with st.expander(label, expanded=False):
                # Session metadata header
                meta_parts = [
                    f'<span><strong>User:</strong> {uname}</span>',
                    f'<span><strong>Session:</strong> <span class="session-card-id">{sid}</span></span>',
                ]
                if time_range:
                    meta_parts.append(f'<span><strong>Time:</strong> {time_range}</span>')
                meta_parts.append(f'<span><strong>Tokens:</strong> ~{format_number(total_tok)}</span>')
                if docs_list:
                    doc_names = ", ".join(docs_list[:3])
                    if len(docs_list) > 3:
                        doc_names += f" +{len(docs_list) - 3} more"
                    meta_parts.append(f'<span><strong>Docs:</strong> {doc_names}</span>')
                st.markdown(
                    f'<div class="stat-meta-row" style="margin-bottom:0.75rem;">'
                    f'{"".join(meta_parts)}'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                # Chat replay
                for m in reversed(msgs):
                    ts_str  = to_local(m["timestamp_utc"]).strftime("%H:%M:%S") if m["timestamp_utc"] else ""
                    dur_str = format_duration(float(m["duration_s"])) if m["duration_s"] else ""
                    tok_str = f"~{format_number(m['tokens_est'])} tok" if m["tokens_est"] else ""
                    with st.chat_message(m["role"]):
                        st.markdown(m["content"])
                        meta_items = [p for p in [ts_str, dur_str, tok_str] if p]
                        if meta_items:
                            st.markdown(
                                f'<div class="msg-meta">'
                                f'{"".join(f"<span>{p}</span>" for p in meta_items)}'
                                f'</div>',
                                unsafe_allow_html=True,
                            )

    # ---- Flat Log tab ----
    with tab_flat:
        for m in rows:
            role      = m["role"]
            role_cls  = "role-user" if role == "user" else "role-assistant"
            role_label = "User" if role == "user" else "Assistant"
            ts_str    = to_local(m["timestamp_utc"]).strftime("%Y-%m-%d %H:%M:%S") if m["timestamp_utc"] else ""
            dur_str   = format_duration(float(m["duration_s"])) if m["duration_s"] else ""
            tok_str   = f"~{format_number(m['tokens_est'])} tok" if m["tokens_est"] else ""
            uname     = m.get("username") or "—"
            preview   = m["content"][:280].replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", " ")
            if len(m["content"]) > 280:
                preview += " …"

            meta_spans = "".join(
                f"<span>{p}</span>" for p in [ts_str, dur_str, tok_str] if p
            )
            st.markdown(
                f'<div class="history-row {role_cls}">'
                f'<div class="history-row-header">'
                f'<div style="display:flex;align-items:center;gap:8px;">'
                f'<span class="history-role-badge {role_cls}">{role_label}</span>'
                f'<span style="font-size:0.78rem;font-weight:600;color:var(--text-primary);">{uname}</span>'
                f'</div>'
                f'<div class="history-row-ident">'
                f'<span class="session-card-id">{m["session_id"][:16]}</span>'
                f'</div>'
                f'</div>'
                f'<div class="history-content">{preview}</div>'
                f'<div class="history-row-meta">{meta_spans}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

    st.markdown('</div>', unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Page entry point
# ---------------------------------------------------------------------------
def main():
    is_admin = "admin" in st.session_state.get("user_roles", {})

    # Code mode is admin-only — force off for non-admins
    if not is_admin and st.session_state.get("code_mode"):
        st.session_state.code_mode = False

    if not st.session_state.chat_active:
        render_lobby()
        if is_admin:
            render_admin()
        return

    render_toolbar()
    render_chat()


main()
